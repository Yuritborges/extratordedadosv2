"""
╔═══════════════════════════════════════════════════════════════════════════════════════╗
║                         COFRE BRASUL - CONSULTOR DE INSUMOS BRASUL                    ║
║                                                                                       ║
║  Sistema de extração automática de códigos FDE a partir de Atestados de obras.        ║
║  Desenvolvido para a Brasul Construtora LTDA.                                         ║
║                                                                                       ║
║  Funcionalidades do programa:                                                         ║
║    - Detecta automaticamente o tipo de página (ACUMULADO DE MEDIÇÃO/QUANTITATIVO)     ║
║    - Extrai códigos FDE via OCR (Tesseract)                                           ║
║    - Corrige rotação de páginas escaneadas                                            ║
║    - Busca descrições e unidades na Base Mestra                                       ║
║    - Gera planilha Excel com todos os itens extraídos                                 ║
║                                                                                       ║
║  Autor: Brasul Construtora LTDA                                                       ║
║  Versão: 2.0                                                                          ║
╚═══════════════════════════════════════════════════════════════════════════════════════╝
"""

import re
import sys
import os
from pathlib import Path

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÕES CENTRALIZADAS
# ═══════════════════════════════════════════════════════════════════════════════
# Aqui puxamos as configurações que ficam no arquivo config/settings.py.
# Isso facilita quando precisamos mudar algum caminho ou ajustar parâmetros,
# pois tudo fica concentrado em um só lugar.
# ═══════════════════════════════════════════════════════════════════════════════

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from config.settings import (
    MODO_PASTA, PASTA_INPUT, PASTA_OUTPUT, CAMINHO_BASE, NOME_SAIDA,
    CAMINHO_PDF, DPI_DETECT, DPI_OCR, CONTRASTE, TESSERACT_CMD
)

# ═══════════════════════════════════════════════════════════════════════════════
# BIBLIOTECAS NECESSÁRIAS
# ═══════════════════════════════════════════════════════════════════════════════
# Aqui fazemos a importação das bibliotecas que o programa precisa.
# O bloco try/except é útil porque se faltar alguma, o programa já avisa
# qual comando rodar para instalar.
# ═══════════════════════════════════════════════════════════════════════════════

try:
    import pytesseract
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
    from PIL import Image, ImageEnhance
    import pdfplumber
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    sys.exit(f"[ERRO] Dependência faltando: {e}  →  pip install {e.name}")

# ═══════════════════════════════════════════════════════════════════════════════
# CORREÇÕES PARA O OCR
# ═══════════════════════════════════════════════════════════════════════════════
# Quando o Tesseract lê um texto digitalizado, ele costuma confundir algumas
# letras com números. Por exemplo, a letra 'O' é lida como '0', o 'Q' também,
# e o 'I' pode virar '1'. A tabela abaixo faz essas correções automaticamente.
# ═══════════════════════════════════════════════════════════════════════════════

_OCR_CORR = str.maketrans('OoQqIlAa', '00000100')

# ═══════════════════════════════════════════════════════════════════════════════
# PADRÃO PARA ENCONTRAR CÓDIGOS FDE
# ═══════════════════════════════════════════════════════════════════════════════
# O padrão abaixo busca por códigos no formato XX.XX.XXX, mas é bem flexível.
# Ele aceita separadores como ponto, traço, underline, barra ou até espaço.
# Também tolera erros comuns do OCR, como letras no lugar de números.
# ═══════════════════════════════════════════════════════════════════════════════

_RE_TOL = re.compile(
    r'(?<![0-9A-Za-z])'                # Garante que não tem caractere antes
    r'([0-2OoAaQq9][0-9OoQq])'         # Primeiro grupo: 2 caracteres (ex: 02, 16)
    r'[.,_|\-\s\\]{0,2}'               # Separador flexível (pode ter ou não)
    r'([0-9OoQq]{2})'                  # Segundo grupo: 2 dígitos (ex: 01, 13)
    r'[.,_|\-\s\\]{0,2}'               # Outro separador flexível
    r'([0-9OoQq]{2,4})'                # Terceiro grupo: 2 a 4 dígitos
    r'(?![0-9A-Za-z])'                 # Garante que não tem caractere depois
)

# ═══════════════════════════════════════════════════════════════════════════════
# CÓDIGOS GLOBAIS (CABEÇALHOS DE GRUPOS)
# ═══════════════════════════════════════════════════════════════════════════════
# Estes são códigos "mãe" que representam categorias inteiras de serviços.
# Por exemplo, 0100000 significa "SERVIÇOS GERAIS". Quando o OCR encontra
# um código específico que não está na Base Mestra, a gente usa esses grupos
# como fallback para não deixar o item sem descrição.
# ═══════════════════════════════════════════════════════════════════════════════

_GRUPOS_GLOBAIS = {
    '0100000': ('SERVIÇOS GERAIS', '%'),
    '0200000': ('FUNDAÇÕES', '%'),
    '0300000': ('ESTRUTURA', '%'),
    '0400000': ('DEMOLIÇÕES', '%'),
    '0500000': ('ESQUADRIAS DE MADEIRA', '%'),
    '0600000': ('ESQUADRIAS METÁLICAS', '%'),
    '0700000': ('COBERTURA', '%'),
    '0800000': ('INST. HIDRÁULICAS E SANITÁRIAS', '%'),
    '0900000': ('INST. ELÉTRICAS E LÓGICA', '%'),
    '1000000': ('FORROS', '%'),
    '1100000': ('IMPERMEABILIZAÇÕES', '%'),
    '1200000': ('REVESTIMENTOS', '%'),
    '1300000': ('PISOS', '%'),
    '1400000': ('VIDROS', '%'),
    '1500000': ('PINTURAS', '%'),
    '1600000': ('SERVIÇOS COMPLEMENTARES', '%'),
    # Subgrupos que aparecem com frequência nos atestados
    '0702000': ('ESTRUTURA DE COBERTURA METALICA', '%'),
    '0704000': ('PECAS PARA COBERTURA', '%'),
    '0905000': ('REDE DE BAIXA TENSAO: DUTO/QUADROS PARCIAIS LUZ/QUADROS TELEFONE', '%'),
    '0907000': ('REDE DE BAIXA TENSAO: ENFIACAO', '%'),
    '0908000': ('PONTOS DE INTERRUPTORES E TOMADAS', '%'),
    '0909000': ('LUMINARIAS INTERNAS', '%'),
    '0913000': ('PARA RAIOS', '%'),
    '0982000': ('CONSERVACAO - BAIXA TENSAO', '%'),
    '0984000': ('CONSERVACAO - APARELHOS E EQUIPAMENTOS', '%'),
}

# ═══════════════════════════════════════════════════════════════════════════════
# CORREÇÕES MANUAIS DE CÓDIGOS
# ═══════════════════════════════════════════════════════════════════════════════
# Quando o OCR lê um código errado, podemos corrigir manualmente.
# Por exemplo, se ele leu '1606058' mas na verdade é '1606059', a gente
# coloca aqui a correção para que o programa acerte na próxima execução.
# ═══════════════════════════════════════════════════════════════════════════════

_OCR_ITEM_FIX = {
    '1606058': '1606059',   # Exemplo de correção de código específico
}

# ═══════════════════════════════════════════════════════════════════════════════
# CORREÇÕES DE SUBGRUPOS
# ═══════════════════════════════════════════════════════════════════════════════
# Às vezes o OCR acerta os últimos dígitos mas erra o grupo principal.
# Por exemplo, '0613' pode ser na verdade '0813' (instalações hidráulicas).
# Essa tabela tenta corrigir esses casos.
# ═══════════════════════════════════════════════════════════════════════════════

_OCR_SUBGRUPO_FIX = {
    '0613': '0813',
    '0609': '0809',
    '0611': '0811',
    '0619': '0819',
    '0819': '0809',
    '0919': '0909',
}


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO: carregar_base
# ═══════════════════════════════════════════════════════════════════════════════
# Carrega o arquivo Excel da Base Mestra, que contém todos os códigos FDE
# com suas respectivas descrições e unidades de medida.
# Retorna dois dicionários para consulta rápida:
#   - por_cod: busca direta pelo código de 7 dígitos
#   - por_desc: lista com códigos e descrições para buscas mais flexíveis
# ═══════════════════════════════════════════════════════════════════════════════

def carregar_base(caminho: Path) -> tuple:
    """Carrega a Base Mestra e retorna dois dicionários para consulta."""

    if not caminho.exists():
        sys.exit(f"\n[ERRO] Base Mestra não encontrada:\n  {caminho}\n")

    wb = openpyxl.load_workbook(str(caminho), data_only=True)
    ws = wb.active

    por_cod = {}      # Dicionário principal: código de 7 dígitos → (descrição, unidade)
    por_desc = []     # Lista auxiliar: (código, descrição em maiúsculo)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        # Remove pontos e traços, deixa só os números
        c7 = re.sub(r'\D', '', str(row[0]))
        if len(c7) != 7:
            continue

        desc = str(row[1] or '').strip()
        un = str(row[2] or '').strip()

        por_cod[c7] = (desc, un)
        por_desc.append((c7, desc.upper()))

    print(f"[OK] Base Mestra carregada: {len(por_cod)} itens")
    return por_cod, por_desc


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO: match_item
# ═══════════════════════════════════════════════════════════════════════════════
# Recebe um código de 7 dígitos (extraído do OCR) e tenta encontrar sua descrição.
# A busca segue esta ordem:
#   1. Procura o código exato na Base Mestra
#   2. Se não achar, tenta corrigir o subgrupo (ex: 0613 vira 0813)
#   3. Se ainda não achar, busca no dicionário de grupos globais
# Se não encontrar nada, retorna o código sem descrição mesmo.
# ═══════════════════════════════════════════════════════════════════════════════

def match_item(cod7: str, por_cod: dict, por_desc: list) -> dict:
    """
    Busca um código na Base Mestra.
    Retorna um dicionário com código formatado, descrição e unidade.
    """

    # Aplica correção manual se existir para este código específico
    if cod7 in _OCR_ITEM_FIX:
        cod7 = _OCR_ITEM_FIX[cod7]

    # Formata o código com pontos para ficar no padrão: XX.XX.XXX
    cod_fmt = f"{cod7[:2]}.{cod7[2:4]}.{cod7[4:]}"

    # Primeira tentativa: código exato na Base Mestra
    if cod7 in por_cod:
        desc, un = por_cod[cod7]
        return {'codigo': cod_fmt, 'descricao': desc, 'unidade': un}

    # Segunda tentativa: corrigir o subgrupo (primeiros 4 dígitos)
    sg = cod7[:4]
    if sg in _OCR_SUBGRUPO_FIX:
        cod7_fix = _OCR_SUBGRUPO_FIX[sg] + cod7[4:]
        if cod7_fix in por_cod:
            desc, un = por_cod[cod7_fix]
            cod_fmt_fix = f"{cod7_fix[:2]}.{cod7_fix[2:4]}.{cod7_fix[4:]}"
            return {'codigo': cod_fmt_fix, 'descricao': desc, 'unidade': un}

    # Terceira tentativa: grupos globais (cabeçalhos de categoria)
    if cod7 in _GRUPOS_GLOBAIS:
        desc, un = _GRUPOS_GLOBAIS[cod7]
        return {'codigo': cod_fmt, 'descricao': desc, 'unidade': un}

    # Se não encontrou nada, retorna o código sem descrição
    return {'codigo': cod_fmt, 'descricao': '', 'unidade': ''}


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO: _melhor_candidato_4dig
# ═══════════════════════════════════════════════════════════════════════════════
# Quando o OCR captura um código com terceiro grupo de 4 dígitos (ex: 09.03.0012),
# a gente precisa descobrir qual dos 4 dígitos está sobrando. Essa função tenta
# diferentes combinações, descartando um dígito de cada vez, e retorna a primeira
# que forma um código válido (7 dígitos, grupo entre 01 e 16).
# ═══════════════════════════════════════════════════════════════════════════════

def _melhor_candidato_4dig(a2: str, b2: str, c4: str, por_cod: dict) -> str | None:
    """Gera candidatos para códigos com terceiro grupo de 4 dígitos."""

    # Lista com as combinações possíveis, descartando um dígito por vez
    cands = [
        a2 + b2 + c4[1:],           # descarta o primeiro dígito
        a2 + b2 + c4[0] + c4[2:],   # descarta o segundo dígito
        a2 + b2 + c4[:2] + c4[3:],  # descarta o terceiro dígito
        a2 + b2 + c4[:3],           # descarta o último dígito
    ]

    # Filtra apenas os códigos com 7 dígitos e grupo principal entre 01 e 16
    validos = [c for c in cands if len(c) == 7 and c.isdigit() and 1 <= int(c[:2]) <= 16]

    if not validos:
        return None

    # Prioriza códigos que existem na Base Mestra
    for c in validos:
        if c in por_cod:
            return c

    # Se nenhum existe na base, retorna o primeiro candidato válido
    return validos[0]


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO: _tem_valor_positivo
# ═══════════════════════════════════════════════════════════════════════════════
# Usada nas páginas do tipo QUANTITATIVA. A gente só quer extrair os códigos
# que têm um valor maior que zero na coluna de quantidades. Essa função
# verifica se o texto contém um número positivo no formato brasileiro (1.234,56)
# ou um número inteiro positivo.
# ═══════════════════════════════════════════════════════════════════════════════

def _tem_valor_positivo(txt: str) -> bool:
    """Verifica se o texto contém um valor numérico positivo."""

    # Remove espaços para facilitar a análise
    txt = re.sub(r'\s+', '', txt)

    # Padrão para números no formato brasileiro: 1.234,56 ou 123,45
    padrao_br = re.compile(r'(\d{1,3}(?:\.\d{3})*|\d+),(\d{1,3})')

    for match in padrao_br.finditer(txt):
        try:
            inteiro = match.group(1).replace('.', '')
            decimal = match.group(2)
            valor = float(f"{inteiro}.{decimal}")
            if valor > 0:
                return True
        except ValueError:
            pass

    # Padrão para números inteiros simples
    for n in re.findall(r'\b([1-9]\d{0,6})\b', txt):
        try:
            if int(n) > 0:
                return True
        except ValueError:
            pass

    return False


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES PARA OCR E IMAGEM
# ═══════════════════════════════════════════════════════════════════════════════
# Aqui estão as funções que preparam as imagens para o OCR:
#   - _detectar_rotacao: descobre se a página está torta
#   - _corrigir_rotacao: endireita a página
#   - _prep_img: prepara a imagem (converte para cinza, corrige rotação)
#   - _ocr: executa o reconhecimento de texto propriamente dito
# ═══════════════════════════════════════════════════════════════════════════════

def _detectar_rotacao(img) -> int:
    """Detecta a rotação da página usando o Tesseract."""
    try:
        w, h = img.size
        # Reduz a imagem para processar mais rápido
        pequena = img.resize((w // 2, h // 2), Image.LANCZOS)
        # O Tesseract consegue detectar rotação com a opção --psm 0
        osd = pytesseract.image_to_osd(pequena, config='--psm 0 -l por', nice=0)
        for linha in osd.splitlines():
            if 'Rotate:' in linha:
                return int(linha.split(':')[1].strip())
    except Exception:
        pass
    return 0


def _corrigir_rotacao(img, angulo: int):
    """Corrige a rotação da imagem."""
    if angulo == 0:
        return img
    return img.rotate(-angulo, expand=True)


def _prep_img(page, dpi: int, auto_rotate: bool = True):
    """Prepara uma página PDF para OCR: converte para imagem e corrige rotação."""
    # Converte a página para imagem em escala de cinza
    img = page.to_image(resolution=dpi).original.convert('L')

    if auto_rotate:
        angulo = _detectar_rotacao(img)
        if angulo != 0:
            print(f"       [AVISO] Rotação detectada: {angulo}° — corrigindo")
            img = _corrigir_rotacao(img, angulo)

    return img


def _ocr(img, config: str = '--psm 6 -l por+eng') -> str:
    """Executa OCR em uma imagem."""
    return pytesseract.image_to_string(img, config=config)


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO: _extrair_codigos
# ═══════════════════════════════════════════════════════════════════════════════
# Função principal para extrair códigos FDE de um texto.
# Ela aplica o padrão regex, faz as correções de OCR, valida os códigos
# (grupo entre 01-16, último grupo até 510) e evita duplicatas.
# Se filtrar_base=True, só retorna códigos que existem na Base Mestra.
# ═══════════════════════════════════════════════════════════════════════════════

def _extrair_codigos(texto: str, por_cod: dict, filtrar_base: bool = False) -> list:
    """
    Extrai códigos FDE de um texto usando regex.
    Se filtrar_base=True, só retorna códigos que existem na Base Mestra.
    """
    vistos = set()      # Usado para evitar duplicatas na mesma página
    codigos = []

    for a, b, c in _RE_TOL.findall(texto):
        # Corrige letras que o OCR confundiu com números
        a2 = a.translate(_OCR_CORR)
        b2 = b.translate(_OCR_CORR)
        c2 = c.translate(_OCR_CORR)

        # Monta o código de 7 dígitos, tratando casos onde o terceiro grupo tem 4 dígitos
        if len(c2) == 4:
            cod7 = _melhor_candidato_4dig(a2, b2, c2, por_cod)
        elif len(c2) == 2:
            raw = a2 + b2 + c2 + '0'
            cod7 = raw if len(raw) == 7 and raw.isdigit() else None
        else:
            raw = a2 + b2 + c2
            cod7 = raw if len(raw) == 7 and raw.isdigit() else None

        if not cod7 or len(cod7) != 7:
            continue

        # Valida se o grupo principal (primeiros 2 dígitos) está entre 01 e 16
        if not (1 <= int(cod7[:2]) <= 16):
            continue

        # Valida se o último grupo (dígitos 5-7) não passa de 510
        if int(cod7[4:]) > 510:
            continue

        # Evita duplicatas na mesma página
        if cod7 in vistos:
            continue

        # Se pediu para filtrar, só aceita códigos que existem na base
        if filtrar_base and cod7 not in por_cod and cod7 not in _GRUPOS_GLOBAIS:
            continue

        vistos.add(cod7)
        codigos.append(cod7)

    return codigos


# ═══════════════════════════════════════════════════════════════════════════════
# EXTRAÇÃO ESPECÍFICA POR TIPO DE PÁGINA
# ═══════════════════════════════════════════════════════════════════════════════
# As próximas funções são especializadas para cada formato de página que
# encontramos nos atestados. Cada uma usa um recorte diferente da imagem
# dependendo de onde os códigos costumam aparecer.
# ═══════════════════════════════════════════════════════════════════════════════

def _extrair_acumulado_por_valor(img, W: int, H: int, por_cod: dict) -> list:
    """
    Extrai códigos de páginas do tipo ACUMULADO.
    Nesse formato, os códigos ficam na coluna esquerda da página.
    """
    H_img = img.size[1]

    # Função auxiliar para fazer OCR numa faixa vertical específica
    def _ocr_faixa(x0, x1, y0=0.22):
        faixa = img.crop((int(W * x0), int(H_img * y0), int(W * x1), int(H_img * 0.96)))
        faixa = faixa.resize((faixa.width * 2, faixa.height * 2), Image.LANCZOS)
        faixa = ImageEnhance.Sharpness(faixa).enhance(2.0)
        return _ocr(faixa)

    # OCR em duas faixas para maior cobertura
    txt = _ocr_faixa(0.03, 0.22) + '\n' + _ocr_faixa(0.03, 0.30)

    return _extrair_codigos(txt, por_cod, filtrar_base=True)


def _extrair_quantitativa_por_valor(img, W: int, H: int, por_cod: dict) -> list:
    """
    Extrai códigos de páginas QUANTITATIVA.
    Aqui a gente só pega os códigos que têm um valor positivo na coluna de quantidades.
    """
    H_img = img.size[1]

    # Região dos códigos (coluna esquerda)
    faixa_cod = img.crop((int(W * 0.02), int(H_img * 0.15), int(W * 0.25), int(H_img * 0.97)))
    faixa_cod = faixa_cod.resize((faixa_cod.width * 2, faixa_cod.height * 2), Image.LANCZOS)
    faixa_cod = ImageEnhance.Sharpness(faixa_cod).enhance(2.0)
    txt_cod = _ocr(faixa_cod, '--psm 6 -l por+eng')

    # Região das quantidades (coluna direita)
    faixa_qtd = img.crop((int(W * 0.45), int(H_img * 0.15), int(W * 0.70), int(H_img * 0.97)))
    faixa_qtd = faixa_qtd.resize((faixa_qtd.width * 2, faixa_qtd.height * 2), Image.LANCZOS)
    faixa_qtd = ImageEnhance.Sharpness(faixa_qtd).enhance(2.0)
    txt_qtd = _ocr(faixa_qtd, '--psm 6 -l por+eng')

    linhas_cod = txt_cod.splitlines()
    linhas_qtd = txt_qtd.splitlines()

    codigos_validos = []
    for i, linha_cod in enumerate(linhas_cod):
        qtd_texto = linhas_qtd[i] if i < len(linhas_qtd) else ''
        if _tem_valor_positivo(qtd_texto):
            codigos_validos.extend(_extrair_codigos(linha_cod, por_cod, filtrar_base=True))

    # Remove duplicatas preservando a ordem
    vistos = set()
    resultado = []
    for c in codigos_validos:
        if c not in vistos:
            vistos.add(c)
            resultado.append(c)

    return resultado


def _extrair_extrato_medicao(page, por_cod: dict) -> list:
    """
    Extrai códigos de páginas EXTRATO.
    Nesse formato, os códigos estão espalhados por toda a página.
    """
    img = _prep_img(page, 400, auto_rotate=True)
    img = ImageEnhance.Contrast(img).enhance(CONTRASTE)
    W, H = img.size

    # OCR em toda a página
    pagina = img.crop((0, 0, W, H))
    pagina = pagina.resize((pagina.width * 2, pagina.height * 2), Image.LANCZOS)
    pagina = ImageEnhance.Sharpness(pagina).enhance(2.0)
    txt_completo = _ocr(pagina, '--psm 6 -l por+eng')

    codigos = _extrair_codigos(txt_completo, por_cod)

    # Remove duplicatas preservando a ordem
    vistos = set()
    resultado = []
    for c in codigos:
        if c not in vistos:
            vistos.add(c)
            resultado.append(c)

    return resultado


def _ocr_meia_pagina(page, por_cod: dict) -> list:
    """
    Extrai códigos da metade esquerda de uma página.
    Usado como fallback quando a página não se encaixa nos padrões conhecidos.
    """
    img = _prep_img(page, 300)
    img = ImageEnhance.Contrast(img).enhance(CONTRASTE)
    W, H = img.size

    # Corta a metade esquerda da página
    metade = img.crop((0, int(H * 0.12), int(W * 0.55), int(H * 0.97)))
    metade = metade.resize((metade.width * 2, metade.height * 2), Image.LANCZOS)
    metade = ImageEnhance.Sharpness(metade).enhance(2.0)

    return _extrair_codigos(_ocr(metade), por_cod)


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO: detectar_tipo
# ═══════════════════════════════════════════════════════════════════════════════
# Esta é uma das funções mais importantes do sistema. Ela analisa o cabeçalho
# da página e decide qual tipo de documento estamos processando.
# Os tipos possíveis são:
#   - EXTRATO: páginas de medição com cabeçalho "EXTRATO DA MEDIÇÃO"
#   - ACUMULADO: páginas com valores acumulados
#   - QUANTITATIVA: páginas com quantitativos de serviços
#   - QUANT_CONTRATO: páginas de contrato
#   - OUTRA: quando não consegue identificar
# ═══════════════════════════════════════════════════════════════════════════════

def detectar_tipo(img) -> str:
    """Detecta o tipo de página analisando o cabeçalho e estrutura."""
    W, H = img.size

    # Analisa o cabeçalho (topo da página)
    cab = img.crop((0, int(H * 0.03), W, int(H * 0.40)))
    cab = ImageEnhance.Contrast(cab).enhance(2.0)
    txt = _ocr(cab, '--psm 3 -l por+eng').upper()

    # Regra 1: Extrato da Medição
    if 'EXTRATO' in txt and ('MEDIÇÃO' in txt or 'MEDICAO' in txt):
        return 'EXTRATO'

    # Regra 2: Páginas de continuação (HIDRÁULICA, PINTURA, QUADRA)
    if ('HIDRÁULICA' in txt or 'PINTURA' in txt or 'QUADRA' in txt):
        col = img.crop((0, int(H * 0.10), int(W * 0.35), int(H * 0.97)))
        col = col.resize((col.width * 2, col.height * 2), Image.LANCZOS)
        txt_col = _ocr(col, '--psm 6 -l por+eng')
        if len(re.findall(r'\b\d{2}[.\-]\d{2}[.\-]\d{3}\b', txt_col)) >= 2:
            return 'EXTRATO'

    # Regra 3: Planilha de Quantitativos
    if 'QUANTITATIVOS' in txt and ('SERVICOS' in txt or 'GLOBAIS' in txt):
        return 'QUANT_CONTRATO'

    # Regra 4: Planilha com cabeçalho "PLANILHA"
    if 'PLANILHA' in txt:
        col = img.crop((0, int(H * 0.15), int(W * 0.35), int(H * 0.97)))
        col = col.resize((col.width * 2, col.height * 2), Image.LANCZOS)
        txt_col = _ocr(col, '--psm 6 -l por+eng')
        if len(re.findall(r'\b\d{2}[.\-]\d{2}[.\-]\d{3}\b', txt_col)) >= 2:
            return 'ACUMULADO'

    # Regra 5: ACUMULADO DE MEDIÇÃO
    if 'ACUMULADO' in txt and ('MEDI' in txt or 'CRITER' in txt):
        return 'ACUMULADO'
    if 'MEDI' in txt and ('UNITARI' in txt or 'CRIT' in txt):
        return 'ACUMULADO'
    if 'CRIT' in txt and ('UNITARI' in txt or 'GLOBAL' in txt):
        return 'ACUMULADO'

    # Regra 6: QUANTITATIVA
    if 'QUANTITATIV' in txt:
        return 'QUANTITATIVA'

    # Regra 7: QUANT_CONTRATO (genérico)
    if 'CONTRATO' in txt and 'COD' in txt:
        return 'QUANT_CONTRATO'

    return 'OUTRA'


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO: processar_pagina
# ═══════════════════════════════════════════════════════════════════════════════
# Processa uma única página do PDF. Ela recebe o tipo detectado e chama a
# função de extração adequada. No final, converte os códigos encontrados
# em itens completos (com descrição e unidade) usando a função match_item.
# ═══════════════════════════════════════════════════════════════════════════════

def processar_pagina(page, tipo: str, por_cod: dict, por_desc: list,
                     img_corrigida=None) -> list:
    """Processa uma página PDF e retorna os itens extraídos."""

    # Prepara a imagem
    if img_corrigida is not None:
        img = img_corrigida
    else:
        img = _prep_img(page, DPI_OCR, auto_rotate=True)

    img = ImageEnhance.Contrast(img).enhance(CONTRASTE)
    W, H = img.size

    # Escolhe o método baseado no tipo de página detectado
    if tipo == 'ACUMULADO':
        codigos = _extrair_acumulado_por_valor(img, W, H, por_cod)
    elif tipo == 'QUANTITATIVA':
        codigos = _extrair_quantitativa_por_valor(img, W, H, por_cod)
    elif tipo == 'EXTRATO':
        codigos = _extrair_extrato_medicao(page, por_cod)
    else:
        codigos = _ocr_meia_pagina(page, por_cod)

    # Converte os códigos em itens completos
    itens = []
    for cod7 in codigos:
        item = match_item(cod7, por_cod, por_desc)
        if item:
            itens.append(item)

    return itens


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO: extrair_nome
# ═══════════════════════════════════════════════════════════════════════════════
# Tenta extrair o nome da obra a partir do cabeçalho do PDF.
# Procura por padrões como "ESCOLA 62110 - EE Profa Ana Luiza" ou
# "NOME INTERV.: ...". Se não encontrar nada, usa o nome do arquivo como fallback.
# ═══════════════════════════════════════════════════════════════════════════════

def extrair_nome(page, nome_arquivo: str = '') -> str:
    """Extrai o nome da obra a partir do cabeçalho do PDF."""

    img = _prep_img(page, DPI_DETECT)
    txt = _ocr(img, '--psm 3 -l por+eng').upper()

    # Palavras que indicam que não é o nome da obra (descartar)
    _BLOQUEADOS = re.compile(
        r'\bUNITARI[AO]\b|\bDERIOUTROS\b|\bDERIOUT\b'
        r'|\bGOVERNO DO ESTADO\b|\bFUNDA[CC][AA]O PARA\b'
        r'|\bDEPARTAMENTO DE\b|\bHTTPS?\b|\bSEIIGESP\b|\bCONTROLADOR\b'
    )

    # Padrões de corte (remove partes indesejadas do texto)
    _CORTE = re.compile(
        r'\s{3,}|\||\bCONTRATO\b|\bFISCAL\b|\bPI\s*:\b'
        r'|\bDIRETORIA\b|\bMUNIC[II]PIO\b|\bPROCESSO\b|\bSEI\b'
    )

    # Padrões de busca para encontrar o nome da obra
    padroes = [
        r'(\d{5,6}\s*[-]\s*(?:EE|EM|EMEF|EMEFM|ETEC|CEI|CIEJA|CEEJA|DER|CRE)\s+[A-Z][A-Z\w\s\.\-]{5,60})',
        r'ESCOLA\s*[:\|]?\s*(\d{5,6}\s*[-]\s*[A-Z][A-Z\w\s\.\-]{5,60})',
        r'NOME\s+INTERV\.?\s*[:\|]?\s*([A-Z][A-Z\w\s\.\-]{5,70})',
        r'PR[EE]DIO\s*[:\|]?\s*\d{5,6}\s*[-]\s*([A-Z][A-Z\w\s\.\-]{5,60})',
    ]

    for pat in padroes:
        m = re.search(pat, txt)
        if not m:
            continue
        nome = m.group(1).strip()
        nome = nome.split('\n')[0].strip()
        nome = _CORTE.split(nome)[0].strip()
        nome = re.sub(r'[\s\.,\-]+$', '', nome)
        nome = re.sub(
            r'\s+(APROVADA?|PI\s*\d{4}|ORC\.?\s*FDE|[AÁ]REA\s*CONS[TR]|CONTRATO|FASE\s*\d|VISTO|REV\.?\s*\d).*$',
            '', nome, flags=re.IGNORECASE
        ).strip()
        if _BLOQUEADOS.search(nome):
            continue
        if 6 < len(nome) < 90:
            return nome.title()

    # Se não encontrou nada, usa o nome do arquivo
    if nome_arquivo:
        base = re.sub(r'\.pdf$', '', nome_arquivo, flags=re.IGNORECASE)
        base = re.sub(r'[_\-]+', ' ', base).strip()
        if 4 < len(base) < 90:
            return base.title()

    return 'OBRA_DESCONHECIDA'


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO: _carregar_descricoes_manuais
# ═══════════════════════════════════════════════════════════════════════════════
# Se já existe um arquivo de saída, carrega as descrições que foram preenchidas
# manualmente pelo usuário. Isso evita que as correções manuais sejam perdidas
# na próxima execução do programa.
# ═══════════════════════════════════════════════════════════════════════════════

def _carregar_descricoes_manuais(pasta: Path) -> dict:
    """Carrega descrições preenchidas manualmente no Excel anterior."""

    cofre = pasta / NOME_SAIDA
    if not cofre.exists():
        return {}

    try:
        wb = openpyxl.load_workbook(str(cofre), data_only=True)
        ws = wb.active
        manuais = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 6:
                continue
            cod = str(row[3] or '').strip()
            desc = str(row[4] or '').strip()
            un = str(row[5] or '').strip()

            if cod and desc:
                c7 = re.sub(r'\D', '', cod)
                if len(c7) == 7:
                    manuais[c7] = {'descricao': desc, 'unidade': un}

        return manuais
    except Exception:
        return {}


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO: processar_pdf
# ═══════════════════════════════════════════════════════════════════════════════
# Abre um arquivo PDF, processa página por página e consolida os resultados.
# Também evita processar o mesmo PDF duas vezes e extrai o nome da obra.
# ═══════════════════════════════════════════════════════════════════════════════

def processar_pdf(caminho: Path, por_cod: dict, por_desc: list,
                  ja_processados=None) -> dict:
    """Processa um arquivo PDF completo."""

    print(f"\n  [OK] Processando: {caminho.name}")

    # Evita processar o mesmo PDF duas vezes
    if ja_processados and caminho.name in ja_processados:
        print(f"     [AVISO] Arquivo já processado, pulando...")
        return None

    # Tenta abrir o PDF
    try:
        pdf = pdfplumber.open(str(caminho))
    except Exception as e:
        print(f"     [ERRO] Não foi possível abrir o PDF: {e}")
        return None

    nome_obra = 'OBRA_DESCONHECIDA'
    itens_por_tipo = {}
    ultimo_tipo = 'OUTRA'

    with pdf:
        for num, page in enumerate(pdf.pages, 1):
            try:
                # Detecta o tipo da página
                img_det = _prep_img(page, DPI_DETECT, auto_rotate=True)
                img_det = ImageEnhance.Contrast(img_det).enhance(2.0)
                tipo = detectar_tipo(img_det)

                # Tenta recuperar tipo de página de continuação
                if tipo == 'OUTRA' and ultimo_tipo in ('ACUMULADO', 'QUANTITATIVA'):
                    W_d, H_d = img_det.size
                    col_det = img_det.crop((int(W_d * 0.02), int(H_d * 0.15), int(W_d * 0.22), int(H_d * 0.97)))
                    col_det = col_det.resize((col_det.width * 2, col_det.height * 2), Image.LANCZOS)
                    txt_col = _ocr(col_det, '--psm 6 -l por+eng')
                    n_cods = len(re.findall(r'\b\d{2}[.\-]\d{2}[.\-]\d{3}\b', txt_col))
                    if n_cods >= 2:
                        tipo = ultimo_tipo

                ultimo_tipo = tipo
                if tipo == 'OUTRA':
                    continue

                # Prepara imagem para OCR
                img_ocr = _prep_img(page, DPI_OCR, auto_rotate=True)

                # Tenta extrair o nome da obra na primeira página
                if nome_obra == 'OBRA_DESCONHECIDA':
                    nome_obra = extrair_nome(page, nome_arquivo=caminho.name)

                # Processa a página
                itens = processar_pagina(page, tipo, por_cod, por_desc, img_corrigida=img_ocr)
                print(f"     PÁGINA {num:2d} [{tipo:13}] → {len(itens)} itens")

                # Filtra itens com descrições que são lixo (cabeçalhos, rodapés, etc)
                _LIXO_DESC = re.compile(
                    r'GOVERNO DO ESTADO|FUNDACAO PARA|DEPARTAMENTO DE'
                    r'|HTTPS?:|SEIIGESP|CONTROLADOR|ATESTADO|CONSTRUTORA'
                    r'|ASSINADO ELETRONICAMENTE|DATA BASE|NRO\.\s*MEDI'
                    r'|DIRETORIA SUL|FISCAL CONSORCIO|EMISSHO|EMISSAO'
                    r'|PI\s*\d{4}|SEI\s*\d', re.IGNORECASE
                )

                # Consolida os itens, agrupando por código e anotando os tipos
                for item in itens:
                    c7 = re.sub(r'\D', '', item['codigo'])
                    if len(c7) != 7 or not (1 <= int(c7[:2]) <= 16):
                        continue
                    desc = item.get('descricao', '')
                    if desc and _LIXO_DESC.search(desc):
                        continue
                    if c7 not in itens_por_tipo:
                        itens_por_tipo[c7] = {'item': item, 'tipos': set()}
                    itens_por_tipo[c7]['tipos'].add(tipo)

            except Exception as e:
                print(f"     [ERRO] Página {num}: {e}")

    # Prepara o resultado final
    resultado = []
    for c7, dados in itens_por_tipo.items():
        item = dados['item']
        tipos = dados['tipos']
        # Se o mesmo código apareceu em mais de um tipo, marca como 'AMBOS'
        tipo_final = 'AMBOS' if len(tipos) > 1 else list(tipos)[0]
        resultado.append({**item, 'tipo': tipo_final})

    return {'nome': nome_obra, 'arq': caminho.name, 'itens': resultado}


# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÕES PARA O EXCEL DE SAÍDA
# ═══════════════════════════════════════════════════════════════════════════════
# Cores usadas para destacar os diferentes tipos de itens na planilha final.
# Cada tipo tem uma cor específica para facilitar a identificação visual.
# ═══════════════════════════════════════════════════════════════════════════════

_CORES = {
    'ACUMULADO': 'D6E4F0',      # Azul claro
    'QUANTITATIVA': 'D5F5E3',   # Verde claro
    'QUANT_CONTRATO': 'FEF9E7', # Amarelo claro
    'EXTRATO': 'E6F0FA',        # Azul bem claro
    'AMBOS': 'F9EBEA',          # Rosa claro
}
_COR_STUB = 'FADBD8'           # Vermelho claro (itens sem descrição)
_COR_HEADER = '2C3E50'         # Azul escuro para o cabeçalho


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO: gerar_excel
# ═══════════════════════════════════════════════════════════════════════════════
# Cria o arquivo Excel com todos os itens extraídos. Aplica cores diferentes
# para cada tipo de página e mantém as descrições que foram preenchidas
# manualmente em execuções anteriores.
# ═══════════════════════════════════════════════════════════════════════════════

def gerar_excel(obras: list, pasta: Path, manuais: dict) -> Path:
    """Gera o arquivo Excel com os resultados."""

    saida = pasta / NOME_SAIDA
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cofre_Brasul'

    # Configura o cabeçalho
    colunas = ['Obra', 'Obra_Arq', 'Tipo', 'Cod', 'Desc', 'UN']
    larguras = [40, 35, 16, 14, 70, 8]
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor=_COR_HEADER)
    thin = Side(style='thin', color='CCCCCC')
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=False)

    ws.append(colunas)
    for i, (col, larg) in enumerate(zip(colunas, larguras), 1):
        cell = ws.cell(1, i)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = borda
        cell.alignment = align_center
        ws.column_dimensions[cell.column_letter].width = larg
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # Adiciona os dados de cada obra
    for obra in obras:
        for item in obra['itens']:
            desc = item['descricao']
            un = item['unidade']
            cod7 = re.sub(r'\D', '', item['codigo'])

            # Se não tem descrição mas tem manual, usa a manual
            if not desc and cod7 in manuais:
                desc = manuais[cod7]['descricao']
                un = manuais[cod7].get('unidade', un)

            tipo = item.get('tipo', '')
            row = [obra['nome'], obra['arq'], tipo, item['codigo'], desc, un]
            ws.append(row)

            # Aplica a cor baseada no tipo
            cor = _COR_STUB if not desc else _CORES.get(tipo, 'FFFFFF')
            fill = PatternFill('solid', fgColor=cor)

            linha = ws.max_row
            for col_idx in range(1, 7):
                cell = ws.cell(linha, col_idx)
                cell.fill = fill
                cell.border = borda
                cell.alignment = align_center if col_idx in (3, 4, 6) else align_left

    ws.auto_filter.ref = ws.dimensions
    wb.save(str(saida))
    print(f"\n[OK] Arquivo gerado: {saida}")
    return saida


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO PRINCIPAL (MAIN)
# ═══════════════════════════════════════════════════════════════════════════════
# Ponto de entrada do programa. Aqui é onde tudo começa:
#   - Carrega a Base Mestra
#   - Lista os PDFs na pasta input
#   - Processa cada PDF
#   - Gera o arquivo Excel com os resultados
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print('\n' + '=' * 60)
    print('  COFRE BRASUL — EXTRATOR FDE v9.19')
    print('  Sistema de extração automática de códigos de obras')
    print('=' * 60)

    # Cria a pasta de saída se não existir
    PASTA_OUTPUT.mkdir(parents=True, exist_ok=True)

    # Carrega a Base Mestra
    por_cod, por_desc = carregar_base(CAMINHO_BASE)

    # Carrega descrições manuais de execuções anteriores
    manuais = _carregar_descricoes_manuais(PASTA_OUTPUT)

    # Lista os PDFs a processar
    if MODO_PASTA:
        pdfs = sorted(PASTA_INPUT.rglob('*.pdf'))
    else:
        pdfs = [CAMINHO_PDF]

    print(f"\n[OK] PDFs encontrados: {len(pdfs)}")

    # Processa cada PDF
    obras = []
    for pdf_path in pdfs:
        resultado = processar_pdf(pdf_path, por_cod, por_desc)
        if resultado and resultado['itens']:
            obras.append(resultado)
            print(f"     → {len(resultado['itens'])} itens únicos")

    # Verifica se houve resultados
    if not obras:
        print("\n[AVISO] Nenhum item foi extraído.")
        return

    # Estatísticas finais
    total_itens = sum(len(o['itens']) for o in obras)
    print(f"\n[OK] Resumo da execução:")
    print(f"     Obras processadas: {len(obras)}")
    print(f"     Itens extraídos: {total_itens}")

    # Gera o arquivo Excel
    saida = gerar_excel(obras, PASTA_OUTPUT, manuais)
    print(f"\n[OK] Processamento concluído!")
    print(f"     Abra o arquivo: {saida}")


# ═══════════════════════════════════════════════════════════════════════════════
# PONTO DE ENTRADA
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    main()