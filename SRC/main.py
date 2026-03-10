"""
main.py — COFRE BRASUL v9.16
==================================================
Extrator automático de insumos FDE de PDFs de atestados.
Formato de saída: Obra | Obra_Arq | Tipo | Cod | Desc | UN

HISTÓRICO:
  v9.0  → nome da obra corrigido, MODO_PASTA=True, rglob
  v9.1  → candidatos 4 dígitos: retorna 1 candidato consultando a Base
  v9.2  → Base Mestra expandida 244→3160 itens
  v9.3  → filtro terceiro grupo > 510; _GRUPOS_GLOBAIS
  v9.4  → _OCR_SUBGRUPO_FIX; _GRUPOS_GLOBAIS expandido
  v9.5  → CAPTURA INTELIGENTE POR VALOR (QtdOrç/QtdAcum/UN)
  v9.6  → DUPLA FAIXA OCR: 3~22% + 3~30% com união filtrada
  v9.14 → quatro faixas OCR cobrindo screenshots SEI (y0=0.15 e 0.22)
  v9.15 → TRÊS CORREÇÕES CRÍTICAS:
          FIX-1: OCR lê "9" como "8" no grupo principal
          FIX-2: filtrar_base gera STUB para subgrupos conhecidos
          FIX-3: _OCR_ITEM_FIX expandido
  v9.16 → QUATRO NOVAS CORREÇÕES OCR (baseadas em análise de 10 PDFs):
          FIX-4: _OCR_SUBGRUPO_FIX: 09.64→09.84 (OCR "84"→"64") CRÍTICO
          FIX-5: _OCR_SUBGRUPO_FIX: 06.03→08.03 (OCR "08"→"06")
          FIX-6: _OCR_SUBGRUPO_FIX: 14.02→11.02 (OCR "11"→"14")
          FIX-7: _OCR_ITEM_FIX: 4 padrões pontuais novos
"""

import re
import sys
from pathlib import Path

# ══════════════════════════════════════════════════════════
#  1.  CONFIGURAÇÃO
# ══════════════════════════════════════════════════════════

BASE_DIR      = Path(r"C:\Users\Iury\Documents\PROJETO EXTRATOR DE DADOS VERSÃO 2")
MODO_PASTA    = True
CAMINHO_PDF   = BASE_DIR / "DATA" / "input" / "Ana Luiza Florence Borges I.pdf"
PASTA_INPUT   = BASE_DIR / "DATA" / "input"
PASTA_OUTPUT  = BASE_DIR / "DATA" / "output"
CAMINHO_BASE  = BASE_DIR / "Base_Mestra_FDE.xlsx"
NOME_SAIDA    = "Cofre_Brasul.xlsx"

DPI_DETECT    = 200
DPI_OCR       = 400
CONTRASTE     = 2.5
TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# ══════════════════════════════════════════════════════════
#  2.  IMPORTS
# ══════════════════════════════════════════════════════════

try:
    import pytesseract
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
    from PIL import Image, ImageEnhance
    import pdfplumber
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    sys.exit(f"Dependência faltando: {e}  →  pip install {e.name}")

# ══════════════════════════════════════════════════════════
#  3.  CONSTANTES OCR / REGEX
# ══════════════════════════════════════════════════════════

_OCR_CORR = str.maketrans('OoQqIlAa', '00000100')

_RE_TOL = re.compile(
    r'(?<![0-9A-Za-z])'
    r'([0-2OoAaQq9][0-9OoQq])'
    r'[.,_|\-\s\\]{0,2}'
    r'([0-9OoQq]{2})'
    r'[.,_|\-\s\\]{0,2}'
    r'([0-9OoQq]{2,4})'
    r'(?![0-9A-Za-z])'
)

# ══════════════════════════════════════════════════════════
#  4.  DICIONÁRIO DE GRUPOS GLOBAIS (.000)
# ══════════════════════════════════════════════════════════

_GRUPOS_GLOBAIS = {
    '0100000': ('SERVIÇOS GERAIS', '%'),
    '0200000': ('FUNDAÇÕES', '%'),
    '0300000': ('ESTRUTURA', '%'),
    '0400000': ('DEMOLIÇÕES', '%'),
    '0500000': ('ESQUADRIAS DE MADEIRA', '%'),
    '0600000': ('ESQUADRIAS METÁLICAS', '%'),
    '0700000': ('COBERTURA', '%'),
    '0800000': ('INST. HIDRÁULICAS E SANITÁRIAS', '%'),
    '0900000': ('INST. ELÉTRICAS E LÓGICA', '%'),
    '1000000': ('FORROS', '%'),
    '1100000': ('IMPERMEABILIZAÇÕES', '%'),
    '1200000': ('REVESTIMENTOS', '%'),
    '1300000': ('PISOS', '%'),
    '1400000': ('VIDROS', '%'),
    '1500000': ('PINTURAS', '%'),
    '1600000': ('SERVIÇOS COMPLEMENTARES', '%'),
    '0102000': ('SERVIÇOS GERAIS - SERVIÇOS INICIAIS', '%'),
    '0202000': ('FUNDAÇÕES - ESTACAS', '%'),
    '0203000': ('FUNDAÇÕES - FORMAS', '%'),
    '0204000': ('FUNDAÇÕES - ARMADURAS', '%'),
    '0205000': ('FUNDAÇÕES - CONCRETO', '%'),
    '0302000': ('ARMADURA', '%'),
    '0304000': ('ESTRUTURA - ESTRUTURA METÁLICA', '%'),
    '0307000': ('ESTRUTURA - MADEIRAMENTO', '%'),
    '0350000': ('ESTRUTURA - DEMOLIÇÕES', '%'),
    '0502000': ('ESQUADRIAS - PORTAS INTERNAS', '%'),
    '0560000': ('ESQUADRIAS - RETIRADAS DE MADEIRA', '%'),
    '0580000': ('ESQUADRIAS - FERRAGENS', '%'),
    '0601000': ('ESQ. METÁLICAS - PORTÕES E GRADES', '%'),
    '0603000': ('ESQ. METÁLICAS - ESCADAS E CORRIMÃOS', '%'),
    '0609000': ('ESQ. METÁLICAS - GRADIS E TELAS', '%'),
    '0611000': ('ESQ. METÁLICAS - PORTÕES DESLIZANTES', '%'),
    '0660000': ('ESQ. METÁLICAS - RETIRADAS', '%'),
    '0682000': ('ESQ. METÁLICAS - FERRAGENS DIVERSAS', '%'),
    '0685000': ('ESQ. METÁLICAS - RETIRADAS DE FERRAGENS', '%'),
    '0603000': ('ESQ. METÁLICAS - GUARDA-CORPOS E CORRIMÃOS', '%'),
    '0702000': ('ESTRUTURA DE COBERTURA METALICA', '%'),
    '0703000': ('COBERTURA', '%'),
    '0704000': ('PECAS PARA COBERTURA', '%'),
    '0705000': ('COBERTURA - FECHAMENTOS E VEDAÇÕES', '%'),
    '0760000': ('COBERTURA - RETIRADAS DE ESTRUTURA', '%'),
    '0770000': ('COBERTURA - RETIRADAS DE TELHAS', '%'),
    '0780000': ('COBERTURA - MADEIRAMENTO', '%'),
    '0805000': ('INST. HIDRÁULICAS - TUBOS DE COBRE', '%'),
    '0807000': ('INST. HIDRÁULICAS - ÁGUA FRIA', '%'),
    '0808000': ('INST. HIDRÁULICAS - ESGOTO', '%'),
    '0809000': ('INST. HIDRÁULICAS - ÁGUA FRIA', '%'),
    '0811000': ('REDE DE AGUAS PLUVIAIS: TUBULACOES', '%'),
    '0812000': ('REDE DE AGUAS PLUVIAIS: DEMAIS SERVICOS', '%'),
    '0813000': ('INST. HIDRÁULICAS - TUBULAÇÕES GERAIS', '%'),
    '0819000': ('INST. HIDRÁULICAS - RETIRADAS DIVERSAS', '%'),
    '0850000': ('INST. HIDRÁULICAS - DEMOLIÇÕES', '%'),
    '0860000': ('INST. HIDRÁULICAS - RETIRADAS DIV.', '%'),
    '0884000': ('INST. HIDRÁULICAS - PEÇAS DE REPOSIÇÃO SANITÁRIA', '%'),
    '0902000': ('INST. ELÉTRICAS - ENTRADA DE ENERGIA', '%'),
    '0905000': ('REDE DE BAIXA TENSAO: DUTO/QUADROS PARCIAIS LUZ/QUADROS TELEFONE', '%'),
    '0907000': ('REDE DE BAIXA TENSAO: ENFIACAO', '%'),
    '0908000': ('PONTOS DE INTERRUPTORES E TOMADAS', '%'),
    '0909000': ('LUMINARIAS INTERNAS', '%'),
    '0910000': ('INST. ELÉTRICAS - CENTRO DE LUZ', '%'),
    '0912000': ('INST. ELÉTRICAS - EXAUSTÃO', '%'),
    '0913000': ('PARA RAIOS', '%'),
    '0919000': ('INST. ELÉTRICAS - RETIRADAS DIVERSAS', '%'),
    '0960000': ('INST. ELÉTRICAS - RETIRADAS DIVERSAS', '%'),
    '0964000': ('INST. ELÉTRICAS - RETIRADAS DE APARELHOS', '%'),
    '0974000': ('INST. ELÉTRICAS - RECOLOCAÇÕES', '%'),
    '0982000': ('CONSERVACAO - BAIXA TENSAO', '%'),
    '0984000': ('CONSERVACAO - APARELHOS E EQUIPAMENTOS', '%'),
    '0985000': ('INST. ELÉTRICAS - INFRAESTRUTURA LÓGICA', '%'),
    '1050000': ('FORROS - DEMOLIÇÕES', '%'),
    '1150000': ('IMPERMEABILIZAÇÕES - DEMOLIÇÕES', '%'),
    '1202000': ('REVESTIMENTOS - ARGAMASSAS', '%'),
    '1207000': ('REVESTIMENTOS - PASTILHAS', '%'),
    '1302000': ('PISOS - CIMENTADOS', '%'),
    '1305000': ('PISOS - CERÂMICOS', '%'),
    '1306000': ('PISOS - SOLEIRAS', '%'),
    '1380000': ('PISOS - DEMOLIÇÕES', '%'),
    '1501000': ('ESTRUTURA', '%'),
    '1502000': ('PINTURAS - PAREDES E TETOS', '%'),
    '1503000': ('ESQUADRIAS', '%'),
    '1550000': ('PINTURAS - REMOÇÕES', '%'),
    '1580000': ('PINTURAS - RETIRADAS E REMOÇÕES DIVERSAS', '%'),
    '1603000': ('SERV. COMPL. - JARDINAGEM', '%'),
    '1606000': ('SERV. COMPL. - INSTALAÇÕES PROVISÓRIAS', '%'),
    '1608000': ('SERV. COMPL. - SINALIZAÇÃO', '%'),
    '1614000': ('SERV. COMPL. - CANTEIRO DE OBRAS', '%'),
    '1618000': ('SERV. COMPL. - LIMPEZA E SERVIÇOS FINAIS', '%'),
    '1620000': ('SERV. COMPL. - ELETRODUTOS E TUBULAÇÕES', '%'),
    '1630000': ('SERV. COMPL. - ANDAIMES E TAPUMES', '%'),
    '1650000': ('SERV. COMPL. - DEMOLIÇÕES', '%'),
    '1680000': ('SERV. COMPL. - SERVIÇOS FINAIS', '%'),
    '1680097': ('CAÇAMBA DE 4M3 PARA RETIRADA DE ENTULHO', 'UN'),
}

# ══════════════════════════════════════════════════════════
#  5.  CORREÇÕES OCR
# ══════════════════════════════════════════════════════════

# FIX-3 (v9.15) + FIX-7 (v9.16): padrões recorrentes de troca no código completo
_OCR_ITEM_FIX = {
    # v9.15 originais
    '1606058': '1606059',   # 16.06.058 → 16.06.059 (OCR: 8→9 no último dígito)
    '0202005': '0202095',   # 02.02.005 → 02.02.095
    '0202009': '0202095',   # 02.02.009 → 02.02.095
    '0580004': '0580001',   # 05.80.004 → 05.80.001
    '0580080': '0580081',   # 05.80.080 → 05.80.081
    '0580087': '0580081',   # 05.80.087 → 05.80.081
    '0760056': '0760066',   # 07.60.056 → 07.60.066
    '0780005': '0780001',   # 07.80.005 → 07.80.001
    '1503010': '1503011',   # 15.03.010 → 15.03.011
    '0984035': '0984003',   # 09.84.035 → 09.84.003
    '0201027': '0201025',   # 02.01.027 → 02.01.025
    # v9.16 novos — confirmados na análise dos 10 PDFs
    '1001045': '1001049',   # 10.01.045 → 10.01.049 (OCR: 9→5 inversão)
    '1503060': '1503061',   # 15.03.060 → 15.03.061 (OCR: perdeu "1" final)
    '0974008': '0974006',   # 09.74.008 → 09.74.006 (OCR: 6→8)
    '0760067': '0760061',   # 07.60.067 → 07.60.061 (OCR: 1→7)
    '0780089': '0780019',   # 07.80.089 → 07.80.019 (OCR: 1→8, 9→8)
    '0808004': '0908004',   # 08.08.004 → 09.08.004 (OCR: grupo 08→09 parcial)
}

# Correções de subgrupo (primeiros 4 dígitos)
_OCR_SUBGRUPO_FIX = {
    # v9.15 originais
    '0613': '0813',   # 06.13 → 08.13  (OCR leu 8 como 6)
    '0609': '0809',   # 06.09 → 08.09
    '0611': '0811',   # 06.11 → 08.11
    '0619': '0819',   # 06.19 → 08.19
    '0819': '0809',   # 08.19 → 08.09
    '0919': '0909',   # 09.19 → 09.09
    # v9.16 NOVOS — confirmados na análise de 10 PDFs
    '0964': '0984',   # FIX-4: 09.64 → 09.84 (OCR "84"→"64") ← CRÍTICO
                      # Exemplo: 09.64.039 → 09.84.039 (conservação aparelhos)
    '0603': '0803',   # FIX-5: 06.03 → 08.03 (OCR "08"→"06")
                      # Exemplo: 06.03.020 → 08.03.020
    '1402': '1102',   # FIX-6: 14.02 → 11.02 (OCR "11"→"14", dois "1" viram "4")
                      # Exemplo: 14.02.027 → 11.02.027
}

# ══════════════════════════════════════════════════════════
#  5b. SUBGRUPOS VÁLIDOS — populado por carregar_base()
# ══════════════════════════════════════════════════════════
_SUBGRUPOS_VALIDOS: set = set()


# ══════════════════════════════════════════════════════════
#  6.  CARREGAMENTO DA BASE MESTRA
# ══════════════════════════════════════════════════════════

def carregar_base(caminho: Path) -> tuple:
    global _SUBGRUPOS_VALIDOS

    if not caminho.exists():
        sys.exit(f"\nBase Mestra não encontrada:\n  {caminho}\n")

    wb = openpyxl.load_workbook(str(caminho), data_only=True)
    ws = wb.active
    por_cod  = {}
    por_desc = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        c7 = re.sub(r'\D', '', str(row[0]))
        if len(c7) != 7:
            continue
        desc = str(row[1] or '').strip()
        un   = str(row[2] or '').strip()
        por_cod[c7]  = (desc, un)
        por_desc.append((c7, desc.upper()))

    # FIX-2: constrói conjunto de subgrupos (4 dígitos) presentes na base
    _SUBGRUPOS_VALIDOS = (
        {c7[:4] for c7 in por_cod} |
        {c7[:4] for c7 in _GRUPOS_GLOBAIS}
    )

    print(f"  Base Mestra: {len(por_cod)} itens | {len(_SUBGRUPOS_VALIDOS)} subgrupos conhecidos")
    return por_cod, por_desc


# ══════════════════════════════════════════════════════════
#  7.  MATCH DE ITEM NA BASE
# ══════════════════════════════════════════════════════════

def match_item(cod7: str, por_cod: dict, por_desc: list) -> dict:
    """
    Busca hierárquica:
      1.   FIX-3+7: correção de código completo (_OCR_ITEM_FIX)
      2.   Base Mestra (exata)
      2.5. FIX-1: 08.xx → 09.xx (OCR lê "9" como "8")
      3.   FIX-4+5+6: correção de subgrupo (_OCR_SUBGRUPO_FIX)
      4.   Grupos globais .000 (_GRUPOS_GLOBAIS)
      5.   Stub
    """
    # FIX-3+7: correção de código completo antes de qualquer lookup
    if cod7 in _OCR_ITEM_FIX:
        cod7 = _OCR_ITEM_FIX[cod7]

    cod_fmt = f"{cod7[:2]}.{cod7[2:4]}.{cod7[4:]}"

    # 1. Base Mestra (exata)
    if cod7 in por_cod:
        desc, un = por_cod[cod7]
        return {'codigo': cod_fmt, 'descricao': desc, 'unidade': un}

    # FIX-1 (v9.15): 08.xx → 09.xx
    if cod7[:2] == '08':
        alt09 = '09' + cod7[2:]
        if alt09 in por_cod:
            desc, un = por_cod[alt09]
            cod_fmt_alt = f"09.{alt09[2:4]}.{alt09[4:]}"
            return {'codigo': cod_fmt_alt, 'descricao': desc, 'unidade': un}
        if alt09 in _GRUPOS_GLOBAIS:
            desc, un = _GRUPOS_GLOBAIS[alt09]
            cod_fmt_alt = f"09.{alt09[2:4]}.{alt09[4:]}"
            return {'codigo': cod_fmt_alt, 'descricao': desc, 'unidade': un}

    # FIX-4+5+6: correção de subgrupo (_OCR_SUBGRUPO_FIX)
    sg = cod7[:4]
    if sg in _OCR_SUBGRUPO_FIX:
        cod7_fix = _OCR_SUBGRUPO_FIX[sg] + cod7[4:]
        if cod7_fix in por_cod:
            desc, un = por_cod[cod7_fix]
            cod_fmt_fix = f"{cod7_fix[:2]}.{cod7_fix[2:4]}.{cod7_fix[4:]}"
            return {'codigo': cod_fmt_fix, 'descricao': desc, 'unidade': un}
        # Subgrupo corrigido existe em _GRUPOS_GLOBAIS?
        sg_fix_000 = _OCR_SUBGRUPO_FIX[sg] + '000'
        if sg_fix_000 in _GRUPOS_GLOBAIS:
            cod7_fix2 = _OCR_SUBGRUPO_FIX[sg] + cod7[4:]
            cod_fmt_fix = f"{cod7_fix2[:2]}.{cod7_fix2[2:4]}.{cod7_fix2[4:]}"
            desc, un = _GRUPOS_GLOBAIS[sg_fix_000]
            return {'codigo': cod_fmt_fix, 'descricao': desc, 'unidade': un}

    # 3. Grupos globais .000
    if cod7 in _GRUPOS_GLOBAIS:
        desc, un = _GRUPOS_GLOBAIS[cod7]
        return {'codigo': cod_fmt, 'descricao': desc, 'unidade': un}

    # 4. Stub — item real mas ausente da base
    return {'codigo': cod_fmt, 'descricao': '', 'unidade': ''}


# ══════════════════════════════════════════════════════════
#  8.  CANDIDATO ÚNICO PARA TERCEIRO GRUPO DE 4 DÍGITOS
# ══════════════════════════════════════════════════════════

def _melhor_candidato_4dig(a2: str, b2: str, c4: str, por_cod: dict) -> str | None:
    cands = [
        a2 + b2 + c4[1:],
        a2 + b2 + c4[0] + c4[2:],
        a2 + b2 + c4[:2] + c4[3:],
        a2 + b2 + c4[:3],
    ]
    validos = [c for c in cands
               if len(c) == 7 and c.isdigit() and 1 <= int(c[:2]) <= 16]
    if not validos:
        return None
    for c in validos:
        if c in por_cod:
            return c
    pos2 = a2 + b2 + c4[:2] + c4[3:]
    return pos2 if pos2 in validos else validos[0]


# ══════════════════════════════════════════════════════════
#  9.  VERIFICAÇÃO DE VALOR (v9.5)
# ══════════════════════════════════════════════════════════

def _tem_valor_positivo(txt: str) -> bool:
    nums = re.findall(r'(\d[\d.]*)[.,](\d{1,3})(?!\d)', txt)
    for inteiro, dec in nums:
        try:
            v = float(inteiro.replace('.', '') + '.' + dec)
            if v > 0.0:
                return True
        except ValueError:
            pass
    for n in re.findall(r'\b([1-9]\d{0,6})\b', txt):
        try:
            if int(n) > 0:
                return True
        except ValueError:
            pass
    return False


# ══════════════════════════════════════════════════════════
#  10.  UTILIDADES OCR
# ══════════════════════════════════════════════════════════

def _corrigir_rotacao(img):
    """Detecta e corrige páginas rotacionadas (90°, 180°, 270°) via Tesseract OSD."""
    try:
        osd = pytesseract.image_to_osd(img, config='--psm 0 -l osd', nice=0)
        for linha in osd.splitlines():
            if 'Rotate:' in linha:
                angulo = int(linha.split(':')[1].strip())
                if angulo == 90:
                    return img.rotate(90, expand=True)
                elif angulo == 180:
                    return img.rotate(180, expand=True)
                elif angulo == 270:
                    return img.rotate(270, expand=True)
    except Exception:
        pass  # Se OSD falhar, retorna imagem original sem travar
    return img


def _prep_img(page, dpi: int):
    img = page.to_image(resolution=dpi).original.convert('L')
    return _corrigir_rotacao(img)


def _ocr(img, config: str = '--psm 6 -l por+eng') -> str:
    return pytesseract.image_to_string(img, config=config)


def _ocr_coluna_coords(img, W: int,
                       x0: float, x1: float,
                       y0: float, y1: float,
                       upscale: int = 3,
                       conf_min: int = 20) -> list:
    y_offset = int(img.size[1] * y0)
    col = img.crop((int(W * x0), int(img.size[1] * y0),
                    int(W * x1), int(img.size[1] * y1)))
    col = col.resize((col.width * upscale, col.height * upscale), Image.LANCZOS)
    col = ImageEnhance.Sharpness(col).enhance(2.0)
    data = pytesseract.image_to_data(col, config='--psm 6 -l por+eng',
                                     output_type=pytesseract.Output.DICT)
    words = []
    for i in range(len(data['text'])):
        txt = data['text'][i].strip()
        if not txt or data['conf'][i] < conf_min:
            continue
        y_abs = y_offset + data['top'][i] // upscale
        words.append({'text': txt, 'y': y_abs})
    return words


# ══════════════════════════════════════════════════════════
#  11.  EXTRAÇÃO DE CÓDIGOS — BASE
# ══════════════════════════════════════════════════════════

def _extrair_codigos(texto: str, por_cod: dict, filtrar_base: bool = False) -> list:
    """
    FIX-2 (v9.15): quando filtrar_base=True, gera STUB para itens com
    subgrupo conhecido em _SUBGRUPOS_VALIDOS em vez de descartar.
    """
    vistos  = set()
    codigos = []

    for a, b, c in _RE_TOL.findall(texto):
        a2 = a.translate(_OCR_CORR)
        b2 = b.translate(_OCR_CORR)
        c2 = c.translate(_OCR_CORR)

        if len(c2) == 4:
            cod7 = _melhor_candidato_4dig(a2, b2, c2, por_cod)
        elif len(c2) == 2:
            raw  = a2 + b2 + c2 + '0'
            cod7 = raw if len(raw) == 7 and raw.isdigit() else None
        else:
            raw  = a2 + b2 + c2
            cod7 = raw if len(raw) == 7 and raw.isdigit() else None

        if not cod7 or len(cod7) != 7:
            continue
        if not (1 <= int(cod7[:2]) <= 16):
            continue
        if int(cod7[4:]) > 510:
            continue
        if cod7 not in vistos:
            if filtrar_base:
                esta_na_base  = cod7 in por_cod or cod7 in _GRUPOS_GLOBAIS
                subgrupo_ok   = cod7[:4] in _SUBGRUPOS_VALIDOS
                # Checar também após correções de subgrupo
                sg_corrigido  = _OCR_SUBGRUPO_FIX.get(cod7[:4], cod7[:4])
                subgrupo_ok   = subgrupo_ok or (sg_corrigido + '000' in _GRUPOS_GLOBAIS)
                if not esta_na_base and not subgrupo_ok:
                    continue
            vistos.add(cod7)
            codigos.append(cod7)

    return codigos


# ══════════════════════════════════════════════════════════
#  12.  EXTRAÇÃO POR VALOR — ACUMULADO
# ══════════════════════════════════════════════════════════

def _extrair_acumulado_por_valor(img, W: int, H: int, por_cod: dict) -> list:
    H_img = img.size[1]

    def _ocr_faixa(x0, x1, y0):
        f = img.crop((int(W * x0), int(H_img * y0), int(W * x1), int(H_img * 0.96)))
        f = f.resize((f.width * 3, f.height * 3), Image.LANCZOS)
        f = ImageEnhance.Sharpness(f).enhance(2.0)
        return _ocr(f)

    txt = ('\n'.join([
        _ocr_faixa(0.03, 0.22, 0.22),
        _ocr_faixa(0.03, 0.30, 0.22),
        _ocr_faixa(0.03, 0.22, 0.15),
        _ocr_faixa(0.03, 0.30, 0.15),
    ]))
    return _extrair_codigos(txt, por_cod, filtrar_base=True)


# ══════════════════════════════════════════════════════════
#  13.  EXTRAÇÃO POR VALOR — QUANTITATIVA
# ══════════════════════════════════════════════════════════

def _extrair_quantitativa_por_valor(img, W: int, H: int, por_cod: dict) -> list:
    H_img = img.size[1]

    def _ocr_faixa(x0, x1, y0):
        f = img.crop((int(W * x0), int(H_img * y0), int(W * x1), int(H_img * 0.96)))
        f = f.resize((f.width * 3, f.height * 3), Image.LANCZOS)
        f = ImageEnhance.Sharpness(f).enhance(2.0)
        return _ocr(f)

    txt = ('\n'.join([
        _ocr_faixa(0.03, 0.22, 0.22),
        _ocr_faixa(0.03, 0.30, 0.22),
        _ocr_faixa(0.03, 0.22, 0.15),
        _ocr_faixa(0.03, 0.30, 0.15),
    ]))
    return _extrair_codigos(txt, por_cod, filtrar_base=True)


# ══════════════════════════════════════════════════════════
#  14.  OCR MEIA PÁGINA (QUANT_CONTRATO)
# ══════════════════════════════════════════════════════════

def _ocr_meia_pagina(page, por_cod: dict) -> list:
    img = _prep_img(page, 300)
    img = ImageEnhance.Contrast(img).enhance(CONTRASTE)
    W, H = img.size
    metade = img.crop((0, int(H * 0.12), int(W * 0.55), int(H * 0.97)))
    metade = metade.resize((metade.width * 2, metade.height * 2), Image.LANCZOS)
    metade = ImageEnhance.Sharpness(metade).enhance(2.0)
    return _extrair_codigos(_ocr(metade), por_cod)


# ══════════════════════════════════════════════════════════
#  15.  DETECÇÃO DO TIPO DE PÁGINA
# ══════════════════════════════════════════════════════════

def detectar_tipo(img) -> str:
    W, H = img.size
    cab  = img.crop((0, int(H * 0.03), W, int(H * 0.40)))
    cab  = ImageEnhance.Contrast(cab).enhance(2.0)
    txt  = _ocr(cab, '--psm 3 -l por+eng').upper()

    if 'ACUMULADO' in txt and ('MEDI' in txt or 'CRITER' in txt):
        return 'ACUMULADO'
    if 'MEDI' in txt and ('UNITARI' in txt or 'CRIT' in txt):
        return 'ACUMULADO'
    if 'CRIT' in txt and ('UNITARI' in txt or 'GLOBAL' in txt):
        return 'ACUMULADO'
    if 'QUANTITATIV' in txt:
        return 'QUANTITATIVA'
    if 'CONTRATO' in txt and 'COD' in txt:
        return 'QUANT_CONTRATO'
    return 'OUTRA'


# ══════════════════════════════════════════════════════════
#  16.  PROCESSAMENTO DE UMA PÁGINA
# ══════════════════════════════════════════════════════════

def processar_pagina(page, tipo: str, por_cod: dict, por_desc: list) -> list:
    img = _prep_img(page, DPI_OCR)
    img = ImageEnhance.Contrast(img).enhance(CONTRASTE)
    W, H = img.size

    if tipo == 'ACUMULADO':
        codigos = _extrair_acumulado_por_valor(img, W, H, por_cod)
    elif tipo == 'QUANTITATIVA':
        codigos = _extrair_quantitativa_por_valor(img, W, H, por_cod)
    else:
        codigos = _ocr_meia_pagina(page, por_cod)

    itens = []
    for cod7 in codigos:
        item = match_item(cod7, por_cod, por_desc)
        if item:
            itens.append(item)
    return itens


# ══════════════════════════════════════════════════════════
#  17.  EXTRAÇÃO DO NOME DA OBRA
# ══════════════════════════════════════════════════════════

def extrair_nome(page) -> str:
    img = _prep_img(page, DPI_DETECT)
    txt = _ocr(img, '--psm 3 -l por+eng').upper()

    padroes = [
        r'ESCOLA\s*[:\|]?\s*(\d{5,6}\s*[-–]\s*[A-Z][\w\s\.\-]+)',
        r'(\d{5,6}\s*[-–]\s*(?:EE|EM|EMEF|EMEFM|ETEC|CEI|CIEJA)\s+[A-Z][\w\s\.\-]{5,60})',
        r'NOME\s+INTERV\.?\s*[:\|]?\s*([A-Z][\w\s\.\-]{5,70})',
        r'PR[EÉ]DIO\s*[:\|]?\s*\d{5,6}\s*[-–]\s*([A-Z][\w\s\.\-]{5,60})',
    ]

    for pat in padroes:
        m = re.search(pat, txt)
        if m:
            nome = m.group(1).strip()
            nome = re.split(r'\s{3,}|\||\n|CONTRATO|FISCAL|PI\s*:|DIRETORIA', nome)[0].strip()
            nome = re.sub(r'[\s\.\,\-]+$', '', nome)
            if 6 < len(nome) < 90:
                return nome.title()

    return 'OBRA_DESCONHECIDA'


# ══════════════════════════════════════════════════════════
#  18.  PRESERVAÇÃO DE PREENCHIMENTOS MANUAIS
# ══════════════════════════════════════════════════════════

def _carregar_descricoes_manuais(pasta: Path) -> dict:
    cofre = pasta / NOME_SAIDA
    if not cofre.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(str(cofre), data_only=True)
        ws = wb.active
        manuais = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 6:
                continue
            cod = str(row[3] or '').strip()
            desc = str(row[4] or '').strip()
            un   = str(row[5] or '').strip()
            if cod and desc:
                c7 = re.sub(r'\D', '', cod)
                if len(c7) == 7:
                    manuais[c7] = {'descricao': desc, 'unidade': un}
        print(f"  Preenchimentos manuais carregados: {len(manuais)}")
        return manuais
    except Exception:
        return {}


# ══════════════════════════════════════════════════════════
#  19.  PROCESSA UM PDF COMPLETO
# ══════════════════════════════════════════════════════════

def processar_pdf(caminho: Path, por_cod: dict, por_desc: list,
                  ja_processados=None) -> dict:
    print(f"\n  >> {caminho.name}")

    if ja_processados and caminho.name in ja_processados:
        print(f"     (já processado, pulando)")
        return None

    try:
        pdf = pdfplumber.open(str(caminho))
    except Exception as e:
        print(f"     ERRO ao abrir PDF: {e}")
        return None

    nome_obra = 'OBRA_DESCONHECIDA'
    itens_por_tipo = {}
    ultimo_tipo = 'OUTRA'

    with pdf:
        for num, page in enumerate(pdf.pages, 1):
            try:
                img_det = _prep_img(page, DPI_DETECT)
                img_det = ImageEnhance.Contrast(img_det).enhance(2.0)
                tipo = detectar_tipo(img_det)

                if tipo == 'OUTRA' and ultimo_tipo in ('ACUMULADO', 'QUANTITATIVA'):
                    W_d, H_d = img_det.size
                    col_det = img_det.crop((int(W_d*0.02), int(H_d*0.15),
                                            int(W_d*0.22), int(H_d*0.97)))
                    col_det = col_det.resize((col_det.width*2, col_det.height*2),
                                             Image.LANCZOS)
                    txt_col = _ocr(col_det, '--psm 6 -l por+eng')
                    n_cods = len(re.findall(
                        r'\b\d{2}[.\-]\d{2}[.\-]\d{3}\b', txt_col))
                    if n_cods >= 2:
                        tipo = ultimo_tipo

                ultimo_tipo = tipo
                if tipo == 'OUTRA':
                    continue

                if nome_obra == 'OBRA_DESCONHECIDA':
                    nome_obra = extrair_nome(page)

                itens = processar_pagina(page, tipo, por_cod, por_desc)
                n_com_desc = sum(1 for i in itens if i.get('descricao'))
                n_stub     = len(itens) - n_com_desc
                stub_info  = f" ({n_stub} stubs)" if n_stub else ""
                print(f"     pág {num:2d} [{tipo:13}]: {len(itens)} itens{stub_info}")

                for item in itens:
                    c7 = re.sub(r'\D', '', item['codigo'])
                    if c7 not in itens_por_tipo:
                        itens_por_tipo[c7] = {'item': item, 'tipos': set()}
                    itens_por_tipo[c7]['tipos'].add(tipo)

            except Exception as e:
                print(f"     pág {num}: ERRO — {e}")

    resultado = []
    for c7, dados in itens_por_tipo.items():
        item  = dados['item']
        tipos = dados['tipos']
        tipo_final = 'AMBOS' if len(tipos) > 1 else list(tipos)[0]
        resultado.append({**item, 'tipo': tipo_final})

    return {
        'nome': nome_obra,
        'arq' : caminho.name,
        'itens': resultado,
    }


# ══════════════════════════════════════════════════════════
#  20.  GERAÇÃO DO EXCEL DE SAÍDA
# ══════════════════════════════════════════════════════════

_CORES = {
    'ACUMULADO':     'D6E4F0',
    'QUANTITATIVA':  'D5F5E3',
    'QUANT_CONTRATO':'FEF9E7',
    'AMBOS':         'F9EBEA',
}
_COR_STUB   = 'FADBD8'
_COR_HEADER = '2C3E50'


def gerar_excel(obras: list, pasta: Path, manuais: dict) -> Path:
    saida = pasta / NOME_SAIDA
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cofre_Brasul'

    colunas = ['Obra', 'Obra_Arq', 'Tipo', 'Cod', 'Desc', 'UN']
    larguras = [40, 35, 16, 14, 70, 8]
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor=_COR_HEADER)
    thin = Side(style='thin', color='CCCCCC')
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    align_left   = Alignment(horizontal='left',   vertical='center', wrap_text=False)

    ws.append(colunas)
    for i, (col, larg) in enumerate(zip(colunas, larguras), 1):
        cell = ws.cell(1, i)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = borda
        cell.alignment = align_center
        ws.column_dimensions[cell.column_letter].width = larg
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    for obra in obras:
        for item in obra['itens']:
            desc = item['descricao']
            un   = item['unidade']
            cod7 = re.sub(r'\D', '', item['codigo'])

            if not desc and cod7 in manuais:
                desc = manuais[cod7]['descricao']
                un   = manuais[cod7].get('unidade', un)

            tipo = item.get('tipo', '')
            row  = [obra['nome'], obra['arq'], tipo, item['codigo'], desc, un]
            ws.append(row)

            cor  = _COR_STUB if not desc else _CORES.get(tipo, 'FFFFFF')
            fill = PatternFill('solid', fgColor=cor)

            linha = ws.max_row
            for col_idx in range(1, 7):
                cell = ws.cell(linha, col_idx)
                cell.fill      = fill
                cell.border    = borda
                cell.alignment = align_center if col_idx in (3, 4, 6) else align_left

    ws.auto_filter.ref = ws.dimensions
    wb.save(str(saida))
    print(f"\n  Cofre salvo em: {saida}")
    return saida


# ══════════════════════════════════════════════════════════
#  21.  MAIN
# ══════════════════════════════════════════════════════════

def _ler_obras_salvas(pasta: Path) -> tuple[list, set]:
    """Lê o Cofre_Brasul.xlsx já existente e retorna (obras, nomes_arq_ja_processados)."""
    saida = pasta / NOME_SAIDA
    if not saida.exists():
        return [], set()
    try:
        wb  = openpyxl.load_workbook(str(saida), data_only=True)
        ws  = wb.active
        # Reconstrói lista de obras a partir das linhas salvas
        from collections import defaultdict, OrderedDict
        obras_map = OrderedDict()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[3]:   # sem código
                continue
            arq  = str(row[1] or '')
            nome = str(row[0] or '')
            if arq not in obras_map:
                obras_map[arq] = {'nome': nome, 'arq': arq, 'itens': []}
            obras_map[arq]['itens'].append({
                'tipo':      str(row[2] or ''),
                'codigo':    str(row[3] or ''),
                'descricao': str(row[4] or ''),
                'unidade':   str(row[5] or ''),
            })
        obras     = list(obras_map.values())
        ja_feitos = set(obras_map.keys())
        return obras, ja_feitos
    except Exception as e:
        print(f"  ⚠  Não foi possível ler o Cofre existente: {e}")
        return [], set()


def main():
    print('\n' + '=' * 60)
    print('  COFRE BRASUL — Extrator FDE v9.16')
    print('=' * 60)

    PASTA_OUTPUT.mkdir(parents=True, exist_ok=True)

    por_cod, por_desc = carregar_base(CAMINHO_BASE)
    manuais = _carregar_descricoes_manuais(PASTA_OUTPUT)

    if MODO_PASTA:
        pdfs = sorted(PASTA_INPUT.rglob('*.pdf'))
    else:
        pdfs = [CAMINHO_PDF]

    # ── CHECKPOINT: ler o que já foi salvo ──────────────────────────────────
    obras, ja_feitos = _ler_obras_salvas(PASTA_OUTPUT)

    if ja_feitos:
        print(f"\n  ✅ Retomando: {len(ja_feitos)} obra(s) já processada(s) no Cofre.")
        pdfs_pendentes = [p for p in pdfs if p.name not in ja_feitos]
    else:
        pdfs_pendentes = pdfs

    total_pdfs    = len(pdfs)
    total_feitos  = len(ja_feitos)
    total_pendente = len(pdfs_pendentes)

    print(f"  PDFs encontrados: {total_pdfs}  |  "
          f"Pendentes: {total_pendente}  |  Já prontos: {total_feitos}")

    if not pdfs_pendentes:
        print("\n  Todos os PDFs já foram processados. Nada a fazer.")
        print(f"  Cofre em: {PASTA_OUTPUT / NOME_SAIDA}")
        return

    # ── LOOP INCREMENTAL: salva após cada PDF ───────────────────────────────
    for idx, pdf_path in enumerate(pdfs_pendentes, 1):
        print(f"\n  [{total_feitos + idx}/{total_pdfs}] >> {pdf_path.name}")
        resultado = processar_pdf(pdf_path, por_cod, por_desc)
        if resultado and resultado['itens']:
            obras.append(resultado)
            n_stubs = sum(1 for i in resultado['itens'] if not i.get('descricao'))
            print(f"     → {len(resultado['itens'])} itens únicos "
                  f"({n_stubs} stubs para preenchimento manual)")

            # Salva incrementalmente após cada PDF processado
            saida = gerar_excel(obras, PASTA_OUTPUT, manuais)
            print(f"     💾 Salvo ({total_feitos + idx}/{total_pdfs})")

    if not obras:
        print("\n  Nenhum item extraído.")
        return

    total_itens = sum(len(o['itens']) for o in obras)
    total_stubs = sum(sum(1 for i in o['itens'] if not i.get('descricao'))
                      for o in obras)
    print(f"\n  Total de obras:  {len(obras)}")
    print(f"  Total de itens:  {total_itens}")
    print(f"  Stubs (manuais): {total_stubs}")

    saida = gerar_excel(obras, PASTA_OUTPUT, manuais)
    print(f"\n  Pronto! Abra: {saida}")


if __name__ == '__main__':
    main()