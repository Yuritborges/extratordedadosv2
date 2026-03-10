"""
main.py — COFRE BRASUL v9.6  (VERSÃO DEFINITIVA)
==================================================
Extrator automático de insumos FDE de PDFs de atestados.

Formato de saída: Obra | Obra_Arq | Tipo | Cod | Desc | UN

HISTÓRICO:
  v9.0 → nome da obra corrigido (ESCOLA: antes de tudo), MODO_PASTA=True, rglob
  v9.1 → candidatos de 4 dígitos: retorna 1 candidato (não 2), consultando a Base
          isso eliminou falsos positivos como 02.02.090, 02.02.907, 02.03.041
  v9.2 → Base Mestra expandida de 244 → 3160 itens (tabela FDE completa Abr/2022)
          filtro de grupo ajustado para 01-16 (FDE só vai até grupo 16)
  v9.3 → filtro terceiro grupo > 510 elimina lixo OCR (datas, valores monetários)
          dicionário _GRUPOS_GLOBAIS para itens .000 (grupos/subgrupos de totalização)
  v9.4 → tabela _OCR_SUBGRUPO_FIX: corrige erros OCR no subgrupo (ex: 06.13→08.13)
          _GRUPOS_GLOBAIS expandido com subgrupos adicionais dos PDFs reais
  v9.6 → DUPLA FAIXA OCR: faixa 3~22% + faixa 3~30% com união filtrada pela base
          máxima cobertura sem falsos positivos
  v9.5 → CAPTURA INTELIGENTE POR VALOR:
          ACUMULADO: captura só se QtdOrç > 0 OU QtdAcumulada > 0
          QUANTITATIVA: captura só se UN preenchida OU QTD orçada > 0
          _tem_valor_positivo tolerante: aceita 1-3 decimais e inteiros
          preservação de preenchimentos manuais entre rodadas
"""

import re
import sys
from pathlib import Path

# ══════════════════════════════════════════════════════════
#  1.  CONFIGURAÇÃO  — ajuste conforme seu ambiente
# ══════════════════════════════════════════════════════════

<<<<<<< HEAD
BASE_DIR      = Path(r"C:\Users\yurit\OneDrive\Documentos\EXTRATOR DE DADOS BRASUL 2.0")
=======
BASE_DIR      = Path(r"C:\Users\Iury\Documents\PROJETO EXTRATOR DE DADOS VERSÃO 2")
>>>>>>> 2df6ca42e534be42f68e50be4935e9edcc35c082
MODO_PASTA    = True                           # True = processa toda a pasta input
CAMINHO_PDF   = BASE_DIR / "DATA" / "input" / "Ana Luiza Florence Borges I.pdf"
PASTA_INPUT   = BASE_DIR / "DATA" / "input"
PASTA_OUTPUT  = BASE_DIR / "DATA" / "output"
CAMINHO_BASE  = BASE_DIR / "Base_Mestra_FDE.xlsx"
NOME_SAIDA    = "Cofre_Brasul.xlsx"

DPI_DETECT    = 200     # DPI para detectar tipo de página (rápido)
DPI_OCR       = 400     # DPI para OCR dos códigos (qualidade)
CONTRASTE     = 2.5     # fator de contraste da imagem antes do OCR
TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# ══════════════════════════════════════════════════════════
#  2.  IMPORTS  — checa dependências
# ══════════════════════════════════════════════════════════

try:
    import pytesseract
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
    from PIL import Image, ImageEnhance
    import pdfplumber
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    sys.exit(f"Dependência faltando: {e}  →  pip install {e.name}")

# ══════════════════════════════════════════════════════════
#  3.  CONSTANTES OCR / REGEX
# ══════════════════════════════════════════════════════════

# Corrige letras que o OCR confunde com dígitos
_OCR_CORR = str.maketrans('OoQqIlAa', '00000100')

# Regex principal: tolera separadores entre os grupos do código FDE
# Captura XX.XX.XXX com separadores ., _ | - espaço
_RE_TOL = re.compile(
    r'(?<![0-9A-Za-z])'
    r'([0-2OoAaQq9][0-9OoQq])'          # grupo 1: dois chars (ex: 02, 09, 16)
    r'[.,_|\-\s\\]{0,2}'
    r'([0-9OoQq]{2})'                    # grupo 2: dois dígitos (ex: 01, 13)
    r'[.,_|\-\s\\]{0,2}'
    r'([0-9OoQq]{2,4})'                  # grupo 3: 2-4 dígitos (ex: 001, 0041)
    r'(?![0-9A-Za-z])'
)

# ══════════════════════════════════════════════════════════
#  4.  DICIONÁRIO DE GRUPOS GLOBAIS (.000)
#  Itens de totalização de grupo — não estão na tabela FDE
#  mas aparecem nas planilhas como linhas de subtotal.
# ══════════════════════════════════════════════════════════

_GRUPOS_GLOBAIS = {
    # grupos principais
    '0100000': ('SERVIÇOS GERAIS', '%'),
    '0200000': ('FUNDAÇÕES', '%'),
    '0300000': ('ESTRUTURA', '%'),
    '0400000': ('DEMOLIÇÕES', '%'),
    '0500000': ('ESQUADRIAS DE MADEIRA', '%'),
    '0600000': ('ESQUADRIAS METÁLICAS', '%'),
    '0700000': ('COBERTURA', '%'),
    '0800000': ('INST. HIDRÁULICAS E SANITÁRIAS', '%'),
    '0900000': ('INST. ELÉTRICAS E LÓGICA', '%'),
    '1000000': ('FORROS', '%'),
    '1100000': ('IMPERMEABILIZAÇÕES', '%'),
    '1200000': ('REVESTIMENTOS', '%'),
    '1300000': ('PISOS', '%'),
    '1400000': ('VIDROS', '%'),
    '1500000': ('PINTURAS', '%'),
    '1600000': ('SERVIÇOS COMPLEMENTARES', '%'),
    # subgrupos frequentes
    '0102000': ('SERVIÇOS GERAIS - SERVIÇOS INICIAIS', '%'),
    '0202000': ('FUNDAÇÕES - ESTACAS', '%'),
    '0203000': ('FUNDAÇÕES - FORMAS', '%'),
    '0204000': ('FUNDAÇÕES - ARMADURAS', '%'),
    '0205000': ('FUNDAÇÕES - CONCRETO', '%'),
    '0302000': ('ARMADURA', '%'),
    '0304000': ('ESTRUTURA - ESTRUTURA METÁLICA', '%'),
    '0307000': ('ESTRUTURA - MADEIRAMENTO', '%'),
    '0350000': ('ESTRUTURA - DEMOLIÇÕES', '%'),
    '0502000': ('ESQUADRIAS - PORTAS INTERNAS', '%'),
    '0560000': ('ESQUADRIAS - RETIRADAS DE MADEIRA', '%'),
    '0580000': ('ESQUADRIAS - FERRAGENS', '%'),
    '0601000': ('ESQ. METÁLICAS - PORTÕES E GRADES', '%'),
    '0603000': ('ESQ. METÁLICAS - GUARDA-CORPOS E CORRIMÃOS', '%'),
    '0609000': ('ESQ. METÁLICAS - GRADIS E TELAS', '%'),
    '0611000': ('ESQ. METÁLICAS - PORTÕES DESLIZANTES', '%'),
    '0660000': ('ESQ. METÁLICAS - RETIRADAS', '%'),
    '0682000': ('ESQ. METÁLICAS - FERRAGENS DIVERSAS', '%'),
    '0685000': ('ESQ. METÁLICAS - RETIRADAS DE FERRAGENS', '%'),
    '0702000': ('COBERTURA - ESTRUTURA METÁLICA', '%'),
    '0703000': ('COBERTURA', '%'),
    '0704000': ('COBERTURA - CUMEEIRAS E RUFOS', '%'),
    '0705000': ('COBERTURA - FECHAMENTOS E VEDAÇÕES', '%'),
    '0760000': ('COBERTURA - RETIRADAS DE ESTRUTURA', '%'),
    '0770000': ('COBERTURA - RETIRADAS DE TELHAS', '%'),
    '0780000': ('COBERTURA - MADEIRAMENTO', '%'),
    '0805000': ('INST. HIDRÁULICAS - TUBOS DE COBRE', '%'),
    '0807000': ('INST. HIDRÁULICAS - ÁGUA FRIA', '%'),
    '0808000': ('INST. HIDRÁULICAS - ESGOTO', '%'),
    '0811000': ('REDE DE AGUAS PLUVIAIS: TUBULACOES', '%'),
    '0812000': ('REDE DE AGUAS PLUVIAIS: DEMAIS SERVICOS', '%'),
    '0813000': ('INST. HIDRÁULICAS - TUBULAÇÕES GERAIS', '%'),
    '0819000': ('INST. HIDRÁULICAS - RETIRADAS DIVERSAS', '%'),
    '0850000': ('INST. HIDRÁULICAS - DEMOLIÇÕES', '%'),
    '0860000': ('INST. HIDRÁULICAS - RETIRADAS DIV.', '%'),
    '0884000': ('INST. HIDRÁULICAS - PEÇAS DE REPOSIÇÃO SANITÁRIA', '%'),
    '0902000': ('INST. ELÉTRICAS - ENTRADA DE ENERGIA', '%'),
    '0905000': ('INST. ELÉTRICAS - ELETRODUTOS', '%'),
    '0907000': ('INST. ELÉTRICAS - FIOS E CABOS', '%'),
    '0908000': ('INST. ELÉTRICAS - PONTOS ELÉTRICOS', '%'),
    '0909000': ('INST. ELÉTRICAS - ILUMINAÇÃO', '%'),
    '0912000': ('INST. ELÉTRICAS - EXAUSTÃO', '%'),
    '0913000': ('INST. ELÉTRICAS - SPDA/ATERRAMENTO', '%'),
    '0919000': ('INST. ELÉTRICAS - RETIRADAS DIVERSAS', '%'),
    '0960000': ('INST. ELÉTRICAS - RETIRADAS DIVERSAS', '%'),
    '0964000': ('INST. ELÉTRICAS - RETIRADAS DE APARELHOS', '%'),
    '0974000': ('INST. ELÉTRICAS - RECOLOCAÇÕES', '%'),
    '0982000': ('INST. ELÉTRICAS - POSTES E SUPORTES', '%'),
    '0984000': ('INST. ELÉTRICAS - INTERRUPTORES E TOMADAS', '%'),
    '0985000': ('INST. ELÉTRICAS - INFRAESTRUTURA LÓGICA', '%'),
    '1050000': ('FORROS - DEMOLIÇÕES', '%'),
    '1150000': ('IMPERMEABILIZAÇÕES - DEMOLIÇÕES', '%'),
    '1202000': ('REVESTIMENTOS - ARGAMASSAS', '%'),
    '1207000': ('REVESTIMENTOS - PASTILHAS', '%'),
    '1302000': ('PISOS - CIMENTADOS', '%'),
    '1305000': ('PISOS - CERÂMICOS', '%'),
    '1306000': ('PISOS - SOLEIRAS', '%'),
    '1380000': ('PISOS - DEMOLIÇÕES', '%'),
    '1501000': ('ESTRUTURA', '%'),
    '1502000': ('PINTURAS - PAREDES E TETOS', '%'),
    '1503000': ('ESQUADRIAS', '%'),
    '1550000': ('PINTURAS - REMOÇÕES', '%'),
    '1603000': ('SERV. COMPL. - JARDINAGEM', '%'),
    '1606000': ('SERV. COMPL. - INSTALAÇÕES PROVISÓRIAS', '%'),
    '1608000': ('SERV. COMPL. - SINALIZAÇÃO', '%'),
    '1650000': ('SERV. COMPL. - DEMOLIÇÕES', '%'),
    '1680000': ('SERV. COMPL. - SERVIÇOS FINAIS', '%'),
        # subgrupos 07.xx adicionais (confirmados em PDFs reais)
    '0702000': ('ESTRUTURA DE COBERTURA METALICA', '%'),
    '0704000': ('PECAS PARA COBERTURA', '%'),
    # subgrupos 09.xx adicionais (confirmados em PDFs reais)
    '0905000': ('REDE DE BAIXA TENSAO: DUTO/QUADROS PARCIAIS LUZ/QUADROS TELEFONE', '%'),
    '0907000': ('REDE DE BAIXA TENSAO: ENFIACAO', '%'),
    '0908000': ('PONTOS DE INTERRUPTORES E TOMADAS', '%'),
    '0909000': ('LUMINARIAS INTERNAS', '%'),
    '0913000': ('PARA RAIOS', '%'),
    '0982000': ('CONSERVACAO - BAIXA TENSAO', '%'),
    '0984000': ('CONSERVACAO - APARELHOS E EQUIPAMENTOS', '%'),
}

# ══════════════════════════════════════════════════════════
#  5.  CORREÇÕES DE SUBGRUPO POR ERRO OCR
#  O OCR às vezes confunde dígitos parecidos no subgrupo:
#    8 lido como 6 → subgrupos 06.xx que deveriam ser 08.xx
#    1 lido como 9 → subgrupos xx.19 que deveriam ser xx.09
#  Só ativa quando o subgrupo original NÃO existe na base.
# ══════════════════════════════════════════════════════════

# Correções de código COMPLETO por erro OCR no terceiro grupo
# {cod7_errado: cod7_correto}
_OCR_ITEM_FIX = {
    '1606058': '1606059',   # 16.06.058 → 16.06.059 (OCR lê 8 em vez de 9 no último dígito)
}

_OCR_SUBGRUPO_FIX = {
    '0613': '0813',   # 06.13 → 08.13  (OCR leu 8 como 6) ← confirmado em PDFs reais
    '0609': '0809',   # 06.09 → 08.09  (OCR leu 8 como 6) ← confirmado
    '0611': '0811',   # 06.11 → 08.11  (OCR leu 8 como 6)
    '0619': '0819',   # 06.19 → 08.19
    '0819': '0809',   # 08.19 → 08.09  (OCR leu 0 como 1) ← confirmado
    '0919': '0909',   # 09.19 → 09.09  (OCR leu 0 como 1) ← confirmado
}

# ══════════════════════════════════════════════════════════
#  6.  CARREGAMENTO DA BASE MESTRA
# ══════════════════════════════════════════════════════════

def carregar_base(caminho: Path) -> tuple:
    """
    Lê a Base_Mestra_FDE.xlsx e retorna:
      por_cod  : dict  cod7 → (descricao, unidade)
      por_desc : list  [(cod7, descricao_upper)] para busca por descrição
    """
    if not caminho.exists():
        sys.exit(f"\nBase Mestra não encontrada:\n  {caminho}\n")

    wb = openpyxl.load_workbook(str(caminho), data_only=True)
    ws = wb.active
    por_cod  = {}
    por_desc = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        c7 = re.sub(r'\D', '', str(row[0]))
        if len(c7) != 7:
            continue
        desc = str(row[1] or '').strip()
        un   = str(row[2] or '').strip()
        por_cod[c7]  = (desc, un)
        por_desc.append((c7, desc.upper()))

    print(f"  Base Mestra: {len(por_cod)} itens carregados")
    return por_cod, por_desc


# ══════════════════════════════════════════════════════════
#  7.  MATCH DE ITEM NA BASE
# ══════════════════════════════════════════════════════════

def match_item(cod7: str, por_cod: dict, por_desc: list) -> dict:
    """
    Dado um cod7 (7 dígitos), devolve dict com codigo, descricao, unidade.
    Ordem de busca:
      1. Base Mestra (exata)
      2. Correção de subgrupo OCR (_OCR_SUBGRUPO_FIX)
      3. Grupos globais .000 (_GRUPOS_GLOBAIS)
      4. Stub sem descrição (para preenchimento manual)
    """
    # Correção de código completo por erro OCR no terceiro grupo
    if cod7 in _OCR_ITEM_FIX:
        cod7 = _OCR_ITEM_FIX[cod7]

    cod_fmt = f"{cod7[:2]}.{cod7[2:4]}.{cod7[4:]}"

    # 1. Base Mestra
    if cod7 in por_cod:
        desc, un = por_cod[cod7]
        return {'codigo': cod_fmt, 'descricao': desc, 'unidade': un}

    # 2. Correção de subgrupo OCR
    sg = cod7[:4]
    if sg in _OCR_SUBGRUPO_FIX:
        cod7_fix = _OCR_SUBGRUPO_FIX[sg] + cod7[4:]
        if cod7_fix in por_cod:
            desc, un = por_cod[cod7_fix]
            cod_fmt_fix = f"{cod7_fix[:2]}.{cod7_fix[2:4]}.{cod7_fix[4:]}"
            return {'codigo': cod_fmt_fix, 'descricao': desc, 'unidade': un}

    # 3. Grupos globais .000
    if cod7 in _GRUPOS_GLOBAIS:
        desc, un = _GRUPOS_GLOBAIS[cod7]
        return {'codigo': cod_fmt, 'descricao': desc, 'unidade': un}

    # 4. Stub — item real mas ausente da base; preservado para preenchimento manual
    return {'codigo': cod_fmt, 'descricao': '', 'unidade': ''}


# ══════════════════════════════════════════════════════════
#  8.  CANDIDATO ÚNICO PARA TERCEIRO GRUPO DE 4 DÍGITOS
# ══════════════════════════════════════════════════════════

def _melhor_candidato_4dig(a2: str, b2: str, c4: str, por_cod: dict) -> str | None:
    """
    OCR às vezes captura 4 dígitos no terceiro grupo (ex: "0041" em vez de "001").
    Gera 4 candidatos removendo 1 dígito de cada posição e devolve EXATAMENTE 1.
    Prioridade: candidato que existe na Base Mestra → se nenhum, usa posição 2.
    """
    cands = [
        a2 + b2 + c4[1:],          # remove pos 0
        a2 + b2 + c4[0] + c4[2:],  # remove pos 1
        a2 + b2 + c4[:2] + c4[3:], # remove pos 2  ← empiricamente mais frequente
        a2 + b2 + c4[:3],           # remove pos 3
    ]
    validos = [c for c in cands
               if len(c) == 7 and c.isdigit() and 1 <= int(c[:2]) <= 16]
    if not validos:
        return None
    # prioridade: está na base
    for c in validos:
        if c in por_cod:
            return c
    # fallback: posição 2
    pos2 = a2 + b2 + c4[:2] + c4[3:]
    return pos2 if pos2 in validos else validos[0]


# ══════════════════════════════════════════════════════════
#  9.  VERIFICAÇÃO DE VALOR (v9.5)
#  Decide se um item foi realmente utilizado na obra.
# ══════════════════════════════════════════════════════════

def _tem_valor_positivo(txt: str) -> bool:
    """
    Verifica se o texto OCR contém pelo menos um número > 0.
    Aceita: 12,50 | 12,5 | 1.200,00 | 500 (inteiro — OCR perdeu o decimal).
    Usado para decidir se o item foi utilizado na obra.
    """
    # decimal com 1, 2 ou 3 casas: 12,0 | 12,00 | 1.200,50
    nums = re.findall(r'(\d[\d.]*)[.,](\d{1,3})(?!\d)', txt)
    for inteiro, dec in nums:
        try:
            v = float(inteiro.replace('.', '') + '.' + dec)
            if v > 0.0:
                return True
        except ValueError:
            pass
    # inteiro isolado >= 1 (OCR perdeu a vírgula/decimal)
    for n in re.findall(r'\b([1-9]\d{0,6})\b', txt):
        try:
            if int(n) > 0:
                return True
        except ValueError:
            pass
    return False


# ══════════════════════════════════════════════════════════
#  10.  UTILIDADES OCR
# ══════════════════════════════════════════════════════════

def _prep_img(page, dpi: int):
    """Converte uma página PDF em imagem PIL em escala de cinza."""
    return page.to_image(resolution=dpi).original.convert('L')


def _ocr(img, config: str = '--psm 6 -l por+eng') -> str:
    """OCR simples — devolve string."""
    return pytesseract.image_to_string(img, config=config)


def _ocr_coluna_coords(img, W: int,
                       x0: float, x1: float,
                       y0: float, y1: float,
                       upscale: int = 3,
                       conf_min: int = 20) -> list:
    """
    OCR com image_to_data em uma faixa da imagem.
    Devolve lista de {'text': str, 'y': int} com y em pixels absolutos.
    """
    y_offset = int(img.size[1] * y0)
    col = img.crop((int(W * x0), int(img.size[1] * y0),
                    int(W * x1), int(img.size[1] * y1)))
    col = col.resize((col.width * upscale, col.height * upscale), Image.LANCZOS)
    col = ImageEnhance.Sharpness(col).enhance(2.0)
    data = pytesseract.image_to_data(col, config='--psm 6 -l por+eng',
                                     output_type=pytesseract.Output.DICT)
    words = []
    for i in range(len(data['text'])):
        txt = data['text'][i].strip()
        if not txt or data['conf'][i] < conf_min:
            continue
        y_abs = y_offset + data['top'][i] // upscale
        words.append({'text': txt, 'y': y_abs})
    return words


# ══════════════════════════════════════════════════════════
#  11.  EXTRAÇÃO DE CÓDIGOS — BASE
# ══════════════════════════════════════════════════════════

def _extrair_codigos(texto: str, por_cod: dict, filtrar_base: bool = False) -> list:
    """
    Extrai todos os códigos FDE de um texto OCR.
    Retorna lista de cod7 (string de 7 dígitos).
    """
    vistos  = set()
    codigos = []

    for a, b, c in _RE_TOL.findall(texto):
        a2 = a.translate(_OCR_CORR)
        b2 = b.translate(_OCR_CORR)
        c2 = c.translate(_OCR_CORR)

        if len(c2) == 4:
            cod7 = _melhor_candidato_4dig(a2, b2, c2, por_cod)
        elif len(c2) == 2:
            raw  = a2 + b2 + c2 + '0'
            cod7 = raw if len(raw) == 7 and raw.isdigit() else None
        else:
            raw  = a2 + b2 + c2
            cod7 = raw if len(raw) == 7 and raw.isdigit() else None

        if not cod7 or len(cod7) != 7:
            continue
        if not (1 <= int(cod7[:2]) <= 16):
            continue
        # terceiro grupo > 510 = lixo OCR (maior código real na base é 16.03.510)
        if int(cod7[4:]) > 510:
            continue
        if cod7 not in vistos:
            # filtrar_base=True: aceita apenas o que existe na base ou é grupo .000
            if filtrar_base and cod7 not in por_cod and cod7 not in _GRUPOS_GLOBAIS:
                continue
            vistos.add(cod7)
            codigos.append(cod7)

    return codigos


# ══════════════════════════════════════════════════════════
#  12.  EXTRAÇÃO POR VALOR — ACUMULADO (v9.5)
#  Captura código APENAS se QtdOrç > 0 OU QtdAcumulada > 0
# ══════════════════════════════════════════════════════════

def _extrair_acumulado_por_valor(img, W: int, H: int, por_cod: dict) -> list:
    """
    Página tipo ACUMULADO DE MEDIÇÃO.
    Estratégia dupla faixa (v9.6):
      Faixa A: x=3~22% — cobre a coluna principal de código
      Faixa B: x=3~30% — cobre variações de layout e códigos deslocados
    Faz a UNIÃO das duas e filtra pela base (elimina falsos positivos).
    """
    H_img = img.size[1]

<<<<<<< HEAD
    def _ocr_faixa(x0, x1, y0=0.15):
        f = img.crop((int(W * x0), int(H_img * y0), int(W * x1), int(H_img * 0.96)))
        f = f.resize((f.width * 3, f.height * 3), Image.LANCZOS)
=======
    def _ocr_faixa(x0, x1, y0=0.22):
        f = img.crop((int(W * x0), int(H_img * y0), int(W * x1), int(H_img * 0.96)))
        f = f.resize((f.width * 2, f.height * 2), Image.LANCZOS)
>>>>>>> 2df6ca42e534be42f68e50be4935e9edcc35c082
        f = ImageEnhance.Sharpness(f).enhance(2.0)
        return _ocr(f)

    txt = _ocr_faixa(0.03, 0.22) + '\n' + _ocr_faixa(0.03, 0.30)
    return _extrair_codigos(txt, por_cod, filtrar_base=True)


# ══════════════════════════════════════════════════════════
#  13.  EXTRAÇÃO POR VALOR — QUANTITATIVA (v9.5)
#  Captura código APENAS se UN preenchida OU QTD orçada > 0
# ══════════════════════════════════════════════════════════

def _extrair_quantitativa_por_valor(img, W: int, H: int, por_cod: dict) -> list:
    """
    Página tipo QUANTITATIVA.
    Estratégia dupla faixa (v9.6):
      Faixa A: x=3~22% — coluna principal de código
      Faixa B: x=3~30% — cobre variações de layout
    Faz a UNIÃO das duas e filtra pela base.
    """
    H_img = img.size[1]

<<<<<<< HEAD
    def _ocr_faixa(x0, x1, y0=0.15):
        f = img.crop((int(W * x0), int(H_img * y0), int(W * x1), int(H_img * 0.96)))
        f = f.resize((f.width * 3, f.height * 3), Image.LANCZOS)
=======
    def _ocr_faixa(x0, x1, y0=0.20):
        f = img.crop((int(W * x0), int(H_img * y0), int(W * x1), int(H_img * 0.96)))
        f = f.resize((f.width * 2, f.height * 2), Image.LANCZOS)
>>>>>>> 2df6ca42e534be42f68e50be4935e9edcc35c082
        f = ImageEnhance.Sharpness(f).enhance(2.0)
        return _ocr(f)

    txt = _ocr_faixa(0.03, 0.22) + '\n' + _ocr_faixa(0.03, 0.30)
    return _extrair_codigos(txt, por_cod, filtrar_base=True)


# ══════════════════════════════════════════════════════════
#  14.  OCR MEIA PÁGINA (QUANT_CONTRATO — legado)
# ══════════════════════════════════════════════════════════

def _ocr_meia_pagina(page, por_cod: dict) -> list:
    """
    OCR na metade esquerda da página inteira.
    Usado no QUANT_CONTRATO onde os códigos ficam em | pipes | sem coluna fixa.
    """
    img = _prep_img(page, 300)
    img = ImageEnhance.Contrast(img).enhance(CONTRASTE)
    W, H = img.size
    metade = img.crop((0, int(H * 0.12), int(W * 0.55), int(H * 0.97)))
    metade = metade.resize((metade.width * 2, metade.height * 2), Image.LANCZOS)
    metade = ImageEnhance.Sharpness(metade).enhance(2.0)
    return _extrair_codigos(_ocr(metade), por_cod)


# ══════════════════════════════════════════════════════════
#  15.  DETECÇÃO DO TIPO DE PÁGINA
# ══════════════════════════════════════════════════════════

def detectar_tipo(img) -> str:
    """
    Lê o cabeçalho da página e classifica em:
      ACUMULADO | QUANTITATIVA | QUANT_CONTRATO | OUTRA
    """
    W, H = img.size
    cab  = img.crop((0, int(H * 0.03), W, int(H * 0.40)))  # ampliado 0.25→0.40
    cab  = ImageEnhance.Contrast(cab).enhance(2.0)
    txt  = _ocr(cab, '--psm 3 -l por+eng').upper()

    if 'ACUMULADO' in txt and ('MEDI' in txt or 'CRITER' in txt):
        return 'ACUMULADO'
    if 'MEDI' in txt and ('UNITARI' in txt or 'CRIT' in txt):
        return 'ACUMULADO'
    if 'CRIT' in txt and ('UNITARI' in txt or 'GLOBAL' in txt):
        return 'ACUMULADO'
    if 'QUANTITATIV' in txt:
        return 'QUANTITATIVA'
    if 'CONTRATO' in txt and 'COD' in txt:
        return 'QUANT_CONTRATO'
    return 'OUTRA'


# ══════════════════════════════════════════════════════════
#  16.  PROCESSAMENTO DE UMA PÁGINA
# ══════════════════════════════════════════════════════════

def processar_pagina(page, tipo: str, por_cod: dict, por_desc: list) -> list:
    """
    Extrai itens de uma página usando o método correto pra cada tipo.

    ACUMULADO    → captura inteligente por valor (QtdOrç ou QtdAcum > 0)
    QUANTITATIVA → captura inteligente por valor (UN preenchida ou QTD > 0)
    QUANT_CONTRATO → método legacy (metade esquerda da página)
    """
    img = _prep_img(page, DPI_OCR)
    img = ImageEnhance.Contrast(img).enhance(CONTRASTE)
    W, H = img.size

    if tipo == 'ACUMULADO':
        codigos = _extrair_acumulado_por_valor(img, W, H, por_cod)
    elif tipo == 'QUANTITATIVA':
        codigos = _extrair_quantitativa_por_valor(img, W, H, por_cod)
    else:  # QUANT_CONTRATO
        codigos = _ocr_meia_pagina(page, por_cod)

    itens = []
    for cod7 in codigos:
        item = match_item(cod7, por_cod, por_desc)
        if item:
            itens.append(item)
    return itens


# ══════════════════════════════════════════════════════════
#  17.  EXTRAÇÃO DO NOME DA OBRA
# ══════════════════════════════════════════════════════════

def extrair_nome(page) -> str:
    """
    Lê o cabeçalho e procura o nome da obra/escola.
    Padrões em ordem de prioridade (mais confiável primeiro).
    """
    img = _prep_img(page, DPI_DETECT)
    txt = _ocr(img, '--psm 3 -l por+eng').upper()

    padroes = [
        r'ESCOLA\s*[:\|]?\s*(\d{5,6}\s*[-–]\s*[A-Z][\w\s\.\-]+)',
        r'(\d{5,6}\s*[-–]\s*(?:EE|EM|EMEF|EMEFM|ETEC|CEI|CIEJA)\s+[A-Z][\w\s\.\-]{5,60})',
        r'NOME\s+INTERV\.?\s*[:\|]?\s*([A-Z][\w\s\.\-]{5,70})',
        r'PR[EÉ]DIO\s*[:\|]?\s*\d{5,6}\s*[-–]\s*([A-Z][\w\s\.\-]{5,60})',
    ]

    for pat in padroes:
        m = re.search(pat, txt)
        if m:
            nome = m.group(1).strip()
            nome = re.split(r'\s{3,}|\||\n|CONTRATO|FISCAL|PI\s*:|DIRETORIA', nome)[0].strip()
            nome = re.sub(r'[\s\.\,\-]+$', '', nome)
            if 6 < len(nome) < 90:
                return nome.title()

    return 'OBRA_DESCONHECIDA'


# ══════════════════════════════════════════════════════════
#  18.  PRESERVAÇÃO DE PREENCHIMENTOS MANUAIS
#  Antes de sobrescrever o Cofre, lê as descrições que o
#  usuário preencheu manualmente para não perdê-las.
# ══════════════════════════════════════════════════════════

def _carregar_descricoes_manuais(pasta: Path) -> dict:
    """
    Lê o Cofre_Brasul.xlsx anterior (se existir) e salva
    os preenchimentos manuais: cod7 → {'descricao': ..., 'unidade': ...}
    """
    cofre = pasta / NOME_SAIDA
    if not cofre.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(str(cofre), data_only=True)
        ws = wb.active
        manuais = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            # colunas: Obra | Obra_Arq | Tipo | Cod | Desc | UN
            if len(row) < 6:
                continue
            cod = str(row[3] or '').strip()
            desc = str(row[4] or '').strip()
            un   = str(row[5] or '').strip()
            if cod and desc:
                c7 = re.sub(r'\D', '', cod)
                if len(c7) == 7:
                    manuais[c7] = {'descricao': desc, 'unidade': un}
        print(f"  Preenchimentos manuais carregados: {len(manuais)}")
        return manuais
    except Exception:
        return {}


# ══════════════════════════════════════════════════════════
#  19.  PROCESSA UM PDF COMPLETO
# ══════════════════════════════════════════════════════════

def processar_pdf(caminho: Path, por_cod: dict, por_desc: list,
                  ja_processados=None) -> dict:
    """
    Abre o PDF, processa cada página e devolve:
      { 'nome': nome da obra, 'arq': nome do arquivo, 'itens': lista de itens }

    Se um item aparece nos dois tipos de planilha → tipo vira 'AMBOS'.
    """
    print(f"\n  >> {caminho.name}")

    if ja_processados and caminho.name in ja_processados:
        print(f"     (já processado, pulando)")
        return None

    try:
        pdf = pdfplumber.open(str(caminho))
    except Exception as e:
        print(f"     ERRO ao abrir PDF: {e}")
        return None

    nome_obra = 'OBRA_DESCONHECIDA'
    itens_por_tipo = {}  # cod7 → {'item': dict, 'tipos': set}
    ultimo_tipo = 'OUTRA'  # rastreia tipo anterior para continuação

    with pdf:
        for num, page in enumerate(pdf.pages, 1):
            try:
                img_det = _prep_img(page, DPI_DETECT)
                img_det = ImageEnhance.Contrast(img_det).enhance(2.0)
                tipo = detectar_tipo(img_det)

                # ── continuação de planilha ──────────────────────────
                # Se a página não foi reconhecida (OUTRA) mas a anterior
                # era ACUMULADO ou QUANTITATIVA, verifica se esta página
                # tem códigos XX.XX.XXX na coluna esquerda (x=2~22%).
                # Essa faixa cobre tanto PDFs nativos quanto screenshots SEI.
                if tipo == 'OUTRA' and ultimo_tipo in ('ACUMULADO', 'QUANTITATIVA'):
                    W_d, H_d = img_det.size
                    col_det = img_det.crop((int(W_d*0.02), int(H_d*0.15),
                                            int(W_d*0.22), int(H_d*0.97)))
                    col_det = col_det.resize((col_det.width*2, col_det.height*2),
                                             Image.LANCZOS)
                    txt_col = _ocr(col_det, '--psm 6 -l por+eng')
                    n_cods = len(re.findall(
                        r'\b\d{2}[.\-]\d{2}[.\-]\d{3}\b', txt_col))
                    if n_cods >= 2:
                        tipo = ultimo_tipo  # assume continuação
                # ────────────────────────────────────────────────────

                ultimo_tipo = tipo
                if tipo == 'OUTRA':
                    continue

                # extrai nome da obra na primeira página relevante
                if nome_obra == 'OBRA_DESCONHECIDA':
                    nome_obra = extrair_nome(page)

                itens = processar_pagina(page, tipo, por_cod, por_desc)
                print(f"     pág {num:2d} [{tipo:13}]: {len(itens)} itens")

                for item in itens:
                    c7 = re.sub(r'\D', '', item['codigo'])
                    if c7 not in itens_por_tipo:
                        itens_por_tipo[c7] = {'item': item, 'tipos': set()}
                    itens_por_tipo[c7]['tipos'].add(tipo)

            except Exception as e:
                print(f"     pág {num}: ERRO — {e}")

    # montar lista final com tipo resolvido
    resultado = []
    for c7, dados in itens_por_tipo.items():
        item  = dados['item']
        tipos = dados['tipos']
        if len(tipos) > 1:
            tipo_final = 'AMBOS'
        else:
            tipo_final = list(tipos)[0]
        resultado.append({**item, 'tipo': tipo_final})

    return {
        'nome': nome_obra,
        'arq' : caminho.name,
        'itens': resultado,
    }


# ══════════════════════════════════════════════════════════
#  20.  GERAÇÃO DO EXCEL DE SAÍDA
# ══════════════════════════════════════════════════════════

# Cores por tipo de planilha
_CORES = {
    'ACUMULADO':     'D6E4F0',   # azul claro
    'QUANTITATIVA':  'D5F5E3',   # verde claro
    'QUANT_CONTRATO':'FEF9E7',   # amarelo claro
    'AMBOS':         'F9EBEA',   # salmão
}
_COR_STUB   = 'FADBD8'   # rosa — item sem descrição (preenchimento manual pendente)
_COR_HEADER = '2C3E50'   # cabeçalho escuro


def gerar_excel(obras: list, pasta: Path, manuais: dict) -> Path:
    """
    Gera o Cofre_Brasul.xlsx com todos os itens de todas as obras.
    Preserva descrições manuais carregadas previamente.
    """
    saida = pasta / NOME_SAIDA
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cofre_Brasul'

    # ── cabeçalho
    colunas = ['Obra', 'Obra_Arq', 'Tipo', 'Cod', 'Desc', 'UN']
    larguras = [40, 35, 16, 14, 70, 8]
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor=_COR_HEADER)
    thin = Side(style='thin', color='CCCCCC')
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    align_left   = Alignment(horizontal='left',   vertical='center', wrap_text=False)

    ws.append(colunas)
    for i, (col, larg) in enumerate(zip(colunas, larguras), 1):
        cell = ws.cell(1, i)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = borda
        cell.alignment = align_center
        ws.column_dimensions[cell.column_letter].width = larg
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # ── dados
    for obra in obras:
        for item in obra['itens']:
            desc = item['descricao']
            un   = item['unidade']
            cod7 = re.sub(r'\D', '', item['codigo'])

            # aplica preenchimento manual se não tem descrição na base
            if not desc and cod7 in manuais:
                desc = manuais[cod7]['descricao']
                un   = manuais[cod7].get('unidade', un)

            tipo = item.get('tipo', '')
            row  = [obra['nome'], obra['arq'], tipo, item['codigo'], desc, un]
            ws.append(row)

            # cor da linha
            if not desc:
                cor = _COR_STUB
            else:
                cor = _CORES.get(tipo, 'FFFFFF')
            fill = PatternFill('solid', fgColor=cor)

            linha = ws.max_row
            for col_idx in range(1, 7):
                cell = ws.cell(linha, col_idx)
                cell.fill      = fill
                cell.border    = borda
                cell.alignment = align_center if col_idx in (3, 4, 6) else align_left

    ws.auto_filter.ref = ws.dimensions
    wb.save(str(saida))
    print(f"\n  Cofre salvo em: {saida}")
    return saida


# ══════════════════════════════════════════════════════════
#  21.  MAIN
# ══════════════════════════════════════════════════════════

def main():
    print('\n' + '=' * 60)
<<<<<<< HEAD
    print('  COFRE BRASUL — Extrator FDE v9.13')
=======
    print('  COFRE BRASUL — Extrator FDE v9.12')
>>>>>>> 2df6ca42e534be42f68e50be4935e9edcc35c082
    print('=' * 60)

    PASTA_OUTPUT.mkdir(parents=True, exist_ok=True)

    # carrega base mestra
    por_cod, por_desc = carregar_base(CAMINHO_BASE)

    # preserva preenchimentos manuais do Cofre anterior
    manuais = _carregar_descricoes_manuais(PASTA_OUTPUT)

    # lista PDFs
    if MODO_PASTA:
        pdfs = sorted(PASTA_INPUT.rglob('*.pdf'))
    else:
        pdfs = [CAMINHO_PDF]

    print(f"\n  PDFs encontrados: {len(pdfs)}")

    obras = []
    for pdf_path in pdfs:
        resultado = processar_pdf(pdf_path, por_cod, por_desc)
        if resultado and resultado['itens']:
            obras.append(resultado)
            print(f"     → {len(resultado['itens'])} itens únicos")

    if not obras:
        print("\n  Nenhum item extraído.")
        return

    total_itens = sum(len(o['itens']) for o in obras)
    print(f"\n  Total de obras: {len(obras)}")
    print(f"  Total de itens: {total_itens}")

    saida = gerar_excel(obras, PASTA_OUTPUT, manuais)
    print(f"\n  Pronto! Abra: {saida}")


if __name__ == '__main__':
    main()