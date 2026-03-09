"""
atualizar_base.py  —  COFRE BRASUL
====================================
O que esse script faz:
  Lê o Cofre_Brasul.xlsx que você preencheu manualmente,
  pega todos os códigos que ganharam descrição e atualiza
  a Base_Mestra_FDE.xlsx com esses novos dados.

Por que isso importa:
  Toda vez que você roda o main.py e preenche descrições vazias,
  você está expandindo o conhecimento do sistema. Rodando esse script,
  essas descrições passam a existir na Base Mestra e da próxima vez
  que o mesmo código aparecer em qualquer PDF, ele já vai sair preenchido.

Resultado:
  - Base_Mestra_FDE.xlsx atualizada com os novos itens
  - Relatório mostrando o que foi adicionado e o que foi ignorado
  - Um arquivo de log salvo em output/

Como usar:
  1. Rode o main.py normalmente (gera Cofre_Brasul.xlsx com stubs vazios)
  2. Abra o Cofre e preencha manualmente os campos Desc e UN vazios
  3. Salve o Cofre
  4. Rode esse script: python atualizar_base.py
  5. Rode o main.py de novo → os itens que você preencheu já vêm descritos!
"""

import re
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────
#  CONFIGURAÇÃO — mesmos caminhos do main.py
# ─────────────────────────────────────────────────────────

BASE_DIR     = Path(r"C:\Users\Iury\Documents\PROJETO EXTRATOR DE DADOS VERSÃO 2")
CAMINHO_BASE = BASE_DIR / "Base_Mestra_FDE.xlsx"
CAMINHO_COFRE = BASE_DIR / "DATA" / "output" / "Cofre_Brasul.xlsx"
PASTA_OUTPUT  = BASE_DIR / "DATA" / "output"

# ─────────────────────────────────────────────────────────


def _cod7(cod: str) -> str:
    """Extrai 7 dígitos do código. Ex: '02.03.001' → '0203001'."""
    return re.sub(r'\D', '', str(cod or ''))


def carregar_base(path: Path) -> dict:
    """Lê a Base Mestra e retorna dicionário cod7 → {cod, desc, un, linha}."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    por_cod = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cod  = str(row[0]).strip() if row[0] else ''
        desc = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        un   = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        c7   = _cod7(cod)
        if len(c7) == 7:
            por_cod[c7] = {'cod': cod, 'desc': desc, 'un': un, 'linha': i}
    return por_cod, wb, ws


def carregar_cofre(path: Path) -> list:
    """
    Lê o Cofre e retorna lista de itens com descrição preenchida
    que podem enriquecer a Base Mestra.
    Formato: [(cod7, cod_fmt, desc, un), ...]
    """
    if not path.exists():
        sys.exit(f"\nCofre não encontrado:\n  {path}\n"
                 f"Rode o main.py primeiro pra gerar o Cofre.\n")

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    itens = []
    vistos = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Cofre: Obra | Obra_Arq | Tipo | Cod | Desc | UN
        if not row or len(row) < 6:
            continue
        cod_raw = str(row[3]).strip() if row[3] else ''
        desc    = str(row[4]).strip() if row[4] else ''
        un      = str(row[5]).strip() if row[5] else ''

        c7 = _cod7(cod_raw)
        if len(c7) != 7:
            continue
        if not desc:  # não tem descrição → não serve pra base
            continue
        if c7 in vistos:  # deduplicação (mesmo código em várias obras)
            continue

        vistos.add(c7)
        itens.append((c7, cod_raw, desc, un))

    return itens


def atualizar_base():
    print()
    print("=" * 60)
    print("  COFRE BRASUL — Atualizar Base Mestra")
    print("=" * 60)

    # verifica arquivos
    if not CAMINHO_BASE.exists():
        sys.exit(f"\nBase Mestra não encontrada:\n  {CAMINHO_BASE}\n")
    if not CAMINHO_COFRE.exists():
        sys.exit(f"\nCofre não encontrado:\n  {CAMINHO_COFRE}\n")

    print(f"\nlendo Base Mestra...  {CAMINHO_BASE.name}")
    por_cod, wb_base, ws_base = carregar_base(CAMINHO_BASE)
    print(f"  {len(por_cod)} itens na base atual")

    print(f"\nlendo Cofre...  {CAMINHO_COFRE.name}")
    itens_cofre = carregar_cofre(CAMINHO_COFRE)
    print(f"  {len(itens_cofre)} itens com descrição no Cofre")

    # analisar o que fazer com cada item do Cofre
    adicionados  = []  # itens novos que entraram na base
    atualizados  = []  # itens que já existiam mas a desc estava vazia
    ignorados    = []  # itens que já tinham descrição na base (mantém a base)
    invalidos    = []  # códigos com problemas

    for c7, cod_fmt, desc_cofre, un_cofre in itens_cofre:
        if c7 not in por_cod:
            # código não existe na base → adiciona como novo
            adicionados.append((c7, cod_fmt, desc_cofre, un_cofre))
        else:
            item_base = por_cod[c7]
            if not item_base['desc']:
                # existe na base mas descrição vazia → atualiza com a do Cofre
                atualizados.append((c7, cod_fmt, desc_cofre, un_cofre, item_base['linha']))
            else:
                # existe na base COM descrição → a base tem prioridade, ignora
                ignorados.append((c7, cod_fmt, desc_cofre, item_base['desc']))

    print(f"\n  novos itens a adicionar:    {len(adicionados)}")
    print(f"  itens a atualizar desc:     {len(atualizados)}")
    print(f"  itens já completos na base: {len(ignorados)} (mantidos)")

    if not adicionados and not atualizados:
        print("\n  nada a fazer — a base já está atualizada!")
        return

    # confirmar antes de salvar
    print("\n--- itens que vão entrar na base ---")
    for c7, cod, desc, un in adicionados[:15]:
        print(f"  + {cod:12}  {un:5}  {desc[:55]}")
    if len(adicionados) > 15:
        print(f"  ... e mais {len(adicionados)-15} itens")
    for c7, cod, desc, un, linha in atualizados[:10]:
        print(f"  ~ {cod:12}  {un:5}  {desc[:55]}  (atualiza linha {linha})")
    if len(atualizados) > 10:
        print(f"  ... e mais {len(atualizados)-10} atualizações")

    # fazer backup antes de salvar
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup = CAMINHO_BASE.parent / f"Base_Mestra_FDE_backup_{ts}.xlsx"
    import shutil
    shutil.copy2(str(CAMINHO_BASE), str(backup))
    print(f"\n  backup salvo: {backup.name}")

    # aplicar atualizações de descrição vazia
    if atualizados:
        wb_edit = openpyxl.load_workbook(str(CAMINHO_BASE))
        ws_edit = wb_edit.active
        for c7, cod, desc, un, linha in atualizados:
            ws_edit.cell(linha, 2).value = desc
            if un:
                ws_edit.cell(linha, 3).value = un
        wb_edit.save(str(CAMINHO_BASE))
        print(f"  {len(atualizados)} descrições atualizadas")

    # adicionar novos itens no final da base
    if adicionados:
        # recarrega pra pegar o estado mais recente
        wb_edit = openpyxl.load_workbook(str(CAMINHO_BASE))
        ws_edit = wb_edit.active

        # estilo pra linhas novas — fundo levemente amarelo pra identificar
        fill_novo = PatternFill("solid", fgColor="FFFACD")
        font_norm = Font(size=10)
        al_e = Alignment(horizontal="left", vertical="center")
        al_c = Alignment(horizontal="center", vertical="center")
        brd  = Border(
            left   = Side(style="thin", color="CCCCCC"),
            right  = Side(style="thin", color="CCCCCC"),
            top    = Side(style="thin", color="CCCCCC"),
            bottom = Side(style="thin", color="CCCCCC"),
        )

        linha_inicio = ws_edit.max_row + 1
        for c7, cod_fmt, desc, un in adicionados:
            # formata o código corretamente (XX.XX.XXX)
            if re.match(r'^\d{7}$', cod_fmt.replace('.', '').replace(' ', '')):
                cod_fmt = f"{c7[:2]}.{c7[2:4]}.{c7[4:]}"
            row_idx = ws_edit.max_row + 1
            ws_edit.append([cod_fmt, desc, un])
            for col in range(1, 4):
                cell = ws_edit.cell(row_idx, col)
                cell.fill      = fill_novo
                cell.font      = font_norm
                cell.border    = brd
                cell.alignment = al_e if col == 2 else al_c

        wb_edit.save(str(CAMINHO_BASE))
        print(f"  {len(adicionados)} itens novos adicionados (linhas {linha_inicio} em diante)")

    # salvar relatório em txt
    log_path = PASTA_OUTPUT / f"log_atualizacao_base_{ts}.txt"
    PASTA_OUTPUT.mkdir(parents=True, exist_ok=True)
    with open(str(log_path), 'w', encoding='utf-8') as f:
        f.write(f"Atualização da Base Mestra — {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
        f.write(f"Cofre utilizado: {CAMINHO_COFRE.name}\n\n")

        f.write(f"=== ADICIONADOS ({len(adicionados)}) ===\n")
        for c7, cod, desc, un in adicionados:
            f.write(f"  {cod:12}  {un:5}  {desc}\n")

        f.write(f"\n=== ATUALIZADOS ({len(atualizados)}) ===\n")
        for c7, cod, desc, un, linha in atualizados:
            f.write(f"  {cod:12}  {un:5}  {desc}\n")

        f.write(f"\n=== IGNORADOS / JÁ NA BASE ({len(ignorados)}) ===\n")
        for c7, cod, desc_cofre, desc_base in ignorados:
            f.write(f"  {cod:12}  base: '{desc_base[:40]}' | cofre: '{desc_cofre[:40]}'\n")

    print(f"\n  log salvo: {log_path.name}")

    # resumo final
    total_base_novo = len(por_cod) + len(adicionados)
    print()
    print("=" * 60)
    print(f"  Base Mestra atualizada!")
    print(f"  antes: {len(por_cod)} itens  →  agora: {total_base_novo} itens")
    print(f"  +{len(adicionados)} novos  |  ~{len(atualizados)} atualizados")
    print("=" * 60)
    print("\n  Próximo passo: rode o main.py de novo.")
    print("  Os itens que você preencheu já vão sair com descrição.\n")


if __name__ == "__main__":
    atualizar_base()
